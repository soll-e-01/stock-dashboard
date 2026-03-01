from __future__ import annotations

import argparse
import json
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]

# ── 스타일 정의 ──
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
LABEL_FONT = Font(bold=True, size=10)
PCT_FONT = Font(italic=True, size=10, color="666666")
SECTION_FONT = Font(bold=True, size=10, color="1F3864")
NO_DATA_FONT = Font(size=9, italic=True, color="AAAAAA")
SUBTITLE_FONT = Font(size=9, color="888888")
LINK_FONT = Font(color="0563C1", underline="single", size=10)

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
STRIPE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_BG = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
UP_FONT = Font(size=10, color="CC0000")
DOWN_FONT = Font(size=10, color="0066CC")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color="2F5496"))

REPORT_CODES = [
    ("11013", "Q1"),
    ("11012", "Q2"),
    ("11014", "Q3"),
    ("11011", "FY"),
]

IS_ACCOUNT_DEFS: list[dict[str, Any]] = [
    {"key": "revenue", "display": "매출액",
     "ids": ["ifrs-full_Revenue"], "keywords": ["매출액", "수익(매출액)", "영업수익"]},
    {"key": "cogs", "display": "매출원가",
     "ids": ["ifrs-full_CostOfSales"], "keywords": ["매출원가"]},
    {"key": "gross_profit", "display": "매출총이익",
     "ids": ["ifrs-full_GrossProfit"], "keywords": ["매출총이익"]},
    {"key": "sga", "display": "판매비와관리비",
     "ids": ["dart_TotalSellingGeneralAdministrativeExpenses"], "keywords": ["판매비와관리비"]},
    {"key": "operating", "display": "영업이익",
     "ids": ["dart_OperatingIncomeLoss"], "keywords": ["영업이익"]},
    {"key": "fin_income", "display": "금융수익",
     "ids": ["ifrs-full_FinanceIncome"], "keywords": ["금융수익"]},
    {"key": "fin_expense", "display": "금융비용",
     "ids": ["ifrs-full_FinanceCosts"], "keywords": ["금융비용"]},
    {"key": "other_income", "display": "기타수익",
     "ids": ["dart_OtherGains"], "keywords": ["기타수익"]},
    {"key": "other_expense", "display": "기타비용",
     "ids": ["dart_OtherLosses"], "keywords": ["기타비용"]},
    {"key": "pretax", "display": "세전이익",
     "ids": ["ifrs-full_ProfitLossBeforeTax"], "keywords": ["법인세비용차감전"]},
    {"key": "tax", "display": "법인세비용",
     "ids": ["ifrs-full_IncomeTaxExpenseContinuingOperations"], "keywords": ["법인세비용", "법인세"]},
    {"key": "net_income", "display": "당기순이익",
     "ids": ["ifrs-full_ProfitLoss"], "keywords": ["당기순이익"]},
    {"key": "controlling", "display": "지배주주순이익",
     "ids": ["ifrs-full_ProfitLossAttributableToOwnersOfParent"], "keywords": ["지배기업"]},
    {"key": "eps", "display": "기본주당이익(원)",
     "ids": ["ifrs-full_BasicEarningsLossPerShare"], "keywords": ["기본주당이익"]},
]

BS_ACCOUNT_DEFS: list[dict[str, Any]] = [
    {"key": "current_assets", "display": "유동자산",
     "ids": ["ifrs-full_CurrentAssets"], "keywords": ["유동자산"]},
    {"key": "non_current_assets", "display": "비유동자산",
     "ids": ["ifrs-full_NoncurrentAssets"], "keywords": ["비유동자산"]},
    {"key": "total_assets", "display": "자산총계",
     "ids": ["ifrs-full_Assets"], "keywords": ["자산총계"]},
    {"key": "current_liabilities", "display": "유동부채",
     "ids": ["ifrs-full_CurrentLiabilities"], "keywords": ["유동부채"]},
    {"key": "non_current_liabilities", "display": "비유동부채",
     "ids": ["ifrs-full_NoncurrentLiabilities"], "keywords": ["비유동부채"]},
    {"key": "total_liabilities", "display": "부채총계",
     "ids": ["ifrs-full_Liabilities"], "keywords": ["부채총계"]},
    {"key": "total_equity", "display": "자본총계",
     "ids": ["ifrs-full_Equity"], "keywords": ["자본총계"]},
    {"key": "retained_earnings", "display": "이익잉여금",
     "ids": ["ifrs-full_RetainedEarnings"], "keywords": ["이익잉여금"]},
]

MARGIN_ROWS = [
    {"after": "revenue", "label": "  YoY(%)", "type": "yoy", "of": "revenue"},
    {"after": "gross_profit", "label": "  GPM(%)", "type": "margin", "of": "gross_profit"},
    {"after": "operating", "label": "  OPM(%)", "type": "margin", "of": "operating"},
    {"after": "operating", "label": "  YoY(%)", "type": "yoy", "of": "operating"},
    {"after": "net_income", "label": "  NPM(%)", "type": "margin", "of": "net_income"},
    {"after": "net_income", "label": "  YoY(%)", "type": "yoy", "of": "net_income"},
]

INVESTOR_TYPES = ["사모펀드", "투자신탁", "연기금", "외국인", "개인"]

DART_REPORT_URL = "https://dart.fss.or.kr/dsaf001/main.do?rcpNo="


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="관심종목 Equity Research Model 엑셀 생성")
    p.add_argument("--config", default="config.dart.json", help="DART 설정 파일")
    p.add_argument("--krx-config", default="config.krx.json", help="KRX 설정 파일 (선택)")
    p.add_argument("--date", default="", help="대상 일자 YYYYMMDD")
    return p.parse_args()


def parse_target_date(raw: str) -> date:
    return datetime.strptime(raw, "%Y%m%d").date() if raw else date.today()


def to_yyyymmdd(d: date) -> str:
    return d.strftime("%Y%m%d")


def to_yymmdd(d: date) -> str:
    return d.strftime("%y%m%d")


def title_date_text(d: date) -> str:
    return f"{d:%Y/%m/%d}({WEEKDAY_KR[d.weekday()]})"


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def ensure_file(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{label} 파일이 없습니다: {path}")


def parse_amount(v: str) -> int:
    s = str(v).strip().replace(",", "")
    if not s or s == "-":
        return 0
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return 0


def to_eok(won: int) -> int:
    return int(round(won / 100_000_000))


def safe_pct(numerator: int | float, denominator: int | float) -> float | None:
    if not denominator:
        return None
    return (numerator / denominator) * 100


# ---------------------------------------------------------------------------
# DART Financial Client
# ---------------------------------------------------------------------------

class DartFinancialClient:
    def __init__(self, api_key: str, timeout_sec: int = 30, delay: float = 0.5):
        self.api_key = api_key
        self.base_url = "https://opendart.fss.or.kr/api"
        self.timeout_sec = timeout_sec
        self.delay = delay
        self.session = requests.Session()

    def fetch_full_statements(
        self, corp_code: str, bsns_year: int, reprt_code: str, fs_div: str = "CFS",
    ) -> list[dict[str, str]]:
        resp = self.session.get(
            f"{self.base_url}/fnlttSinglAcntAll.json",
            params={
                "crtfc_key": self.api_key,
                "corp_code": corp_code,
                "bsns_year": str(bsns_year),
                "reprt_code": reprt_code,
                "fs_div": fs_div,
            },
            timeout=self.timeout_sec,
        )
        resp.raise_for_status()
        body = resp.json()
        status = body.get("status", "")
        if status == "013":
            return []
        if status != "000":
            return []
        time.sleep(self.delay)
        return body.get("list", [])

    def fetch_disclosures(
        self, corp_code: str, bgn_de: str, end_de: str, page_count: int = 10,
    ) -> list[dict[str, str]]:
        try:
            resp = self.session.get(
                f"{self.base_url}/list.json",
                params={
                    "crtfc_key": self.api_key,
                    "corp_code": corp_code,
                    "bgn_de": bgn_de,
                    "end_de": end_de,
                    "page_count": str(page_count),
                    "sort": "date",
                    "sort_mth": "desc",
                },
                timeout=self.timeout_sec,
            )
            resp.raise_for_status()
            body = resp.json()
            if body.get("status") != "000":
                return []
            time.sleep(self.delay)
            return body.get("list", [])
        except Exception:
            return []


# ---------------------------------------------------------------------------
# Financial Model
# ---------------------------------------------------------------------------

@dataclass
class PeriodLabel:
    year: int
    quarter: str  # "Q1", "Q2", "Q3", "Q4", "FY"
    reprt_code: str

    @property
    def display(self) -> str:
        if self.quarter == "FY":
            return f"FY{self.year}"
        return f"{self.quarter}'{str(self.year)[2:]}"


@dataclass
class FinancialModel:
    corp_name: str
    stock_code: str = ""
    periods: list[PeriodLabel] = field(default_factory=list)
    is_data: dict[str, dict[str, int]] = field(default_factory=dict)
    bs_data: dict[str, dict[str, int]] = field(default_factory=dict)


def match_account(item: dict[str, str], account_def: dict[str, Any]) -> bool:
    aid = item.get("account_id", "")
    for target_id in account_def.get("ids", []):
        if aid == target_id:
            return True
    anm = item.get("account_nm", "")
    for kw in account_def.get("keywords", []):
        if kw in anm:
            return True
    return False


def build_financial_model(
    client: DartFinancialClient,
    corp_code: str,
    corp_name: str,
    stock_code: str,
    years: list[int],
    fs_div: str = "CFS",
) -> FinancialModel:
    model = FinancialModel(corp_name=corp_name, stock_code=stock_code)
    raw_by_period: dict[str, list[dict[str, str]]] = {}
    used_fs_div = fs_div

    for year in years:
        for reprt_code, q_label in REPORT_CODES:
            items = client.fetch_full_statements(corp_code, year, reprt_code, used_fs_div)
            is_items = [x for x in items if x.get("sj_div") in ("IS", "CIS")]
            bs_items = [x for x in items if x.get("sj_div") == "BS"]
            # CFS 데이터 없으면 OFS로 폴백
            if not is_items and used_fs_div == "CFS":
                items = client.fetch_full_statements(corp_code, year, reprt_code, "OFS")
                is_items = [x for x in items if x.get("sj_div") in ("IS", "CIS")]
                bs_items = [x for x in items if x.get("sj_div") == "BS"]
                if is_items:
                    used_fs_div = "OFS"
                    print(f"  → {corp_name}: CFS 없음, OFS(개별) 사용")
            period_key = f"{year}_{q_label}"
            raw_by_period[period_key] = items  # 전체 저장 (IS+BS)
            pl = PeriodLabel(year=year, quarter=q_label, reprt_code=reprt_code)
            if q_label != "FY":
                model.periods.append(pl)
            else:
                q4_pl = PeriodLabel(year=year, quarter="Q4", reprt_code="Q4_calc")
                model.periods.append(q4_pl)
                model.periods.append(pl)

    # IS 데이터 파싱
    for acct_def in IS_ACCOUNT_DEFS:
        key = acct_def["key"]
        model.is_data[key] = {}
        for year in years:
            for reprt_code, q_label in REPORT_CODES:
                period_key = f"{year}_{q_label}"
                all_items = raw_by_period.get(period_key, [])
                is_items = [x for x in all_items if x.get("sj_div") in ("IS", "CIS")]
                value = 0
                for item in is_items:
                    if match_account(item, acct_def):
                        value = parse_amount(item.get("thstrm_amount", "0"))
                        break
                if q_label != "FY":
                    model.is_data[key][f"{year}_{q_label}"] = value
                else:
                    model.is_data[key][f"{year}_FY"] = value
                    if value:
                        q1 = model.is_data[key].get(f"{year}_Q1", 0)
                        q2 = model.is_data[key].get(f"{year}_Q2", 0)
                        q3 = model.is_data[key].get(f"{year}_Q3", 0)
                        q4 = value - q1 - q2 - q3
                        model.is_data[key][f"{year}_Q4"] = q4
                    else:
                        model.is_data[key][f"{year}_Q4"] = 0

    # BS 데이터 파싱 (잔액이므로 Q4 계산 불필요)
    for acct_def in BS_ACCOUNT_DEFS:
        key = acct_def["key"]
        model.bs_data[key] = {}
        for year in years:
            for reprt_code, q_label in REPORT_CODES:
                period_key = f"{year}_{q_label}"
                all_items = raw_by_period.get(period_key, [])
                bs_items = [x for x in all_items if x.get("sj_div") == "BS"]
                value = 0
                for item in bs_items:
                    if match_account(item, acct_def):
                        value = parse_amount(item.get("thstrm_amount", "0"))
                        break
                if q_label != "FY":
                    model.bs_data[key][f"{year}_{q_label}"] = value
                else:
                    model.bs_data[key][f"{year}_FY"] = value
                    # BS는 잔액이므로 Q4 = FY와 동일
                    model.bs_data[key][f"{year}_Q4"] = value

    return model


def get_is_value(model: FinancialModel, key: str, period: PeriodLabel) -> int:
    col_key = f"{period.year}_{period.quarter}"
    return model.is_data.get(key, {}).get(col_key, 0)


def get_bs_value(model: FinancialModel, key: str, period: PeriodLabel) -> int:
    col_key = f"{period.year}_{period.quarter}"
    return model.bs_data.get(key, {}).get(col_key, 0)


def get_yoy_value(model: FinancialModel, key: str, period: PeriodLabel) -> float | None:
    current = get_is_value(model, key, period)
    if not current:
        return None
    prev_year = period.year - 1
    prev_key = f"{prev_year}_{period.quarter}"
    prev_val = model.is_data.get(key, {}).get(prev_key, 0)
    if not prev_val:
        return None
    return ((current - prev_val) / abs(prev_val)) * 100


def calc_ttm(model: FinancialModel, key: str) -> int:
    """최근 4분기 합산 (TTM: Trailing Twelve Months)."""
    quarters_only = [p for p in model.periods if p.quarter not in ("FY",)]
    recent_4 = quarters_only[-4:] if len(quarters_only) >= 4 else quarters_only
    return sum(get_is_value(model, key, p) for p in recent_4)


def calc_valuations(
    model: FinancialModel, price_row: dict | None,
) -> dict[str, float | None]:
    """PER, PBR, PSR, EV/EBITDA 계산."""
    result: dict[str, float | None] = {
        "PER": None, "PBR": None, "PSR": None, "EV/EBITDA": None,
    }
    if not price_row:
        return result

    mktcap_won = parse_amount(price_row.get("MKTCAP", price_row.get("시가총액", price_row.get("market_cap", "0"))))
    if mktcap_won <= 0:
        return result

    # PER = 시가총액 / 지배주주순이익(TTM)
    controlling_ttm = calc_ttm(model, "controlling")
    if not controlling_ttm:
        controlling_ttm = calc_ttm(model, "net_income")
    if controlling_ttm > 0:
        result["PER"] = mktcap_won / controlling_ttm

    # PBR = 시가총액 / 자본총계 (최근)
    recent_equity = _get_latest_bs(model, "total_equity")
    if recent_equity > 0:
        result["PBR"] = mktcap_won / recent_equity

    # PSR = 시가총액 / 매출액(TTM)
    revenue_ttm = calc_ttm(model, "revenue")
    if revenue_ttm > 0:
        result["PSR"] = mktcap_won / revenue_ttm

    # EV/EBITDA = (시가총액 + 부채총계) / 영업이익(TTM) (간이 계산)
    total_liab = _get_latest_bs(model, "total_liabilities")
    operating_ttm = calc_ttm(model, "operating")
    if operating_ttm > 0:
        ev = mktcap_won + total_liab
        result["EV/EBITDA"] = ev / operating_ttm

    return result


def _get_latest_bs(model: FinancialModel, key: str) -> int:
    """BS 데이터에서 가장 최근 값 반환."""
    for p in reversed(model.periods):
        val = get_bs_value(model, key, p)
        if val:
            return val
    return 0


def fetch_segment_data(
    client: DartFinancialClient,
    corp_code: str,
    bsns_year: int,
    reprt_code: str = "11011",
    fs_div: str = "CFS",
) -> list[dict[str, Any]] | None:
    """DART API에서 사업부문별 매출 시도. 실패시 None 반환."""
    try:
        items = client.fetch_full_statements(corp_code, bsns_year, reprt_code, fs_div)
        if not items and fs_div == "CFS":
            items = client.fetch_full_statements(corp_code, bsns_year, reprt_code, "OFS")

        # sj_div로 세그먼트 항목 탐색
        seg_items = [x for x in items if x.get("sj_div") in ("ASIS",)]
        if not seg_items:
            # account_nm에서 '매출' 포함된 세그먼트 탐색
            seg_items = [
                x for x in items
                if x.get("sj_nm", "").startswith("부문")
                and "매출" in x.get("account_nm", "")
            ]

        if not seg_items:
            return None

        segments: list[dict[str, Any]] = []
        total_seg_rev = 0
        for item in seg_items:
            name = item.get("account_nm", "").strip()
            rev = parse_amount(item.get("thstrm_amount", "0"))
            if rev > 0 and name:
                segments.append({"name": name, "revenue": rev})
                total_seg_rev += rev

        if not segments:
            return None

        # 비중 계산
        for seg in segments:
            seg["pct"] = (seg["revenue"] / total_seg_rev * 100) if total_seg_rev else 0

        return sorted(segments, key=lambda x: x["revenue"], reverse=True)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# KRX Data (Optional)
# ---------------------------------------------------------------------------

def try_fetch_krx_data(
    krx_config_path: Path,
    yyyymmdd: str,
    watchlist_codes: set[str],
) -> tuple[list[dict], dict[str, list[dict]]]:
    price_rows: list[dict] = []
    supply_by_type: dict[str, list[dict]] = {}

    if not krx_config_path.exists():
        print("[KRX] config 파일 없음 - 시세/수급 시트 생략")
        return price_rows, supply_by_type

    try:
        cfg = load_json(krx_config_path)
    except Exception:
        print("[KRX] config 로드 실패 - 시세/수급 시트 생략")
        return price_rows, supply_by_type

    krx = cfg.get("krx", {})
    json_url = krx.get("json_url", "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd")
    headers = krx.get("headers", {})
    timeout = int(krx.get("timeout_sec", 30))

    session = requests.Session()
    session.headers.update(headers)

    # Price data
    try:
        price_params = {
            "bld": "dbms/MDC/STAT/standard/MDCSTAT01501",
            "locale": "ko_KR", "mktId": "ALL", "trdDd": yyyymmdd,
            "share": "1", "money": "1", "csvxls_isNo": "false",
        }
        resp = session.post(json_url, data=price_params, timeout=timeout)
        resp.raise_for_status()
        data = resp.json()
        all_prices = data.get("OutBlock_1", data.get("output", []))
        if isinstance(all_prices, list):
            for row in all_prices:
                code = str(row.get("ISU_SRT_CD", row.get("종목코드", ""))).strip()
                if code in watchlist_codes:
                    price_rows.append(row)
        print(f"[KRX] 시세 {len(price_rows)}건 수집")
    except Exception as e:
        print(f"[KRX] 시세 수집 실패: {e}")

    # Supply data
    inv_codes = {"사모펀드": "3100", "투자신탁": "3000", "연기금": "6000", "외국인": "9000", "개인": "8000"}
    for inv_name, inv_code in inv_codes.items():
        try:
            supply_params = {
                "bld": "dbms/MDC/STAT/standard/MDCSTAT02401",
                "locale": "ko_KR", "mktId": "ALL", "invstTpCd": inv_code,
                "strtDd": yyyymmdd, "endDd": yyyymmdd,
                "share": "1", "money": "1", "csvxls_isNo": "false",
            }
            resp = session.post(json_url, data=supply_params, timeout=timeout)
            resp.raise_for_status()
            data = resp.json()
            all_rows = data.get("OutBlock_1", data.get("output", []))
            filtered = []
            if isinstance(all_rows, list):
                for row in all_rows:
                    code = str(row.get("ISU_SRT_CD", row.get("종목코드", ""))).strip()
                    if code in watchlist_codes:
                        filtered.append(row)
            supply_by_type[inv_name] = filtered
            time.sleep(0.5)
        except Exception as e:
            print(f"[KRX] {inv_name} 수급 수집 실패: {e}")
            supply_by_type[inv_name] = []

    if supply_by_type:
        total = sum(len(v) for v in supply_by_type.values())
        print(f"[KRX] 수급 {total}건 수집")

    return price_rows, supply_by_type


# ---------------------------------------------------------------------------
# Excel Writers
# ---------------------------------------------------------------------------

def set_header_row(ws: Any, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _write_section_header(ws: Any, row: int, title: str, num_cols: int) -> None:
    """섹션 구분 헤더 (배경색 채움)."""
    cell = ws.cell(row, 1, title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_BG
    for ci in range(2, num_cols + 1):
        ws.cell(row, ci).fill = SECTION_BG


def write_sheet_equity(
    ws: Any,
    model: FinancialModel,
    target_date: date | None = None,
    price_row: dict | None = None,
    supply_data: dict[str, int] | None = None,
    disclosures: list[dict[str, str]] | None = None,
    segments: list[dict[str, Any]] | None = None,
) -> None:
    num_period_cols = 1 + len(model.periods)

    # ── 타이틀 ──
    ws.cell(1, 1, model.corp_name).font = TITLE_FONT
    ws.cell(1, 1).border = BOTTOM_BORDER
    ws.cell(2, 1, "단위: 억원 (주당이익·밸류에이션 제외)").font = SUBTITLE_FONT

    cursor = 3  # 현재 행 추적

    # ── ■ 시세 개요 ──
    _write_section_header(ws, cursor, "■ 시세 개요", 7)
    cursor += 1
    price_headers = ["종가", "등락(원)", "등락률(%)", "거래량", "거래대금(억)", "시가총액(억)"]
    for ci, h in enumerate(price_headers, 1):
        cell = ws.cell(cursor, ci, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    cursor += 1
    if price_row:
        close = parse_amount(price_row.get("TDD_CLSPRC", price_row.get("종가", "0")))
        change = parse_amount(price_row.get("CMPPREVDD_PRC", price_row.get("대비", "0")))
        pct_str = str(price_row.get("FLUC_RT", price_row.get("등락률", "0"))).replace(",", "").replace("%", "")
        try:
            pct_val = float(pct_str)
        except ValueError:
            pct_val = 0.0
        volume = parse_amount(price_row.get("ACC_TRDVOL", price_row.get("거래량", "0")))
        trdval = to_eok(parse_amount(price_row.get("ACC_TRDVAL", price_row.get("거래대금", "0"))))
        mktcap = to_eok(parse_amount(price_row.get("MKTCAP", price_row.get("시가총액", "0"))))
        price_vals = [close, change, round(pct_val, 2), volume, trdval, mktcap]
        color_font = UP_FONT if change > 0 else DOWN_FONT if change < 0 else None
        for ci, v in enumerate(price_vals, 1):
            cell = ws.cell(cursor, ci, v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right")
            if ci in (2, 3) and color_font:
                cell.font = color_font
            if ci != 3:
                cell.number_format = "#,##0"
    else:
        ws.cell(cursor, 1, "KRX 데이터 없음").font = NO_DATA_FONT
    cursor += 1

    # ── ■ 밸류에이션 ──
    _write_section_header(ws, cursor, "■ 밸류에이션", 6)
    cursor += 1
    val_headers = ["PER(배)", "PBR(배)", "PSR(배)", "EV/EBITDA(배)"]
    for ci, h in enumerate(val_headers, 1):
        cell = ws.cell(cursor, ci, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    cursor += 1
    valuations = calc_valuations(model, price_row)
    val_keys = ["PER", "PBR", "PSR", "EV/EBITDA"]
    if any(valuations[k] is not None for k in val_keys):
        for ci, k in enumerate(val_keys, 1):
            val = valuations[k]
            cell = ws.cell(cursor, ci, round(val, 1) if val is not None else "")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right")
            if val is not None:
                cell.number_format = "0.0"
    else:
        ws.cell(cursor, 1, "데이터 부족").font = NO_DATA_FONT
    cursor += 1

    # ── ■ 투자자별 수급 ──
    _write_section_header(ws, cursor, "■ 투자자별 수급 (순매수, 백만원)", 6)
    cursor += 1
    for ci, inv in enumerate(INVESTOR_TYPES, 1):
        cell = ws.cell(cursor, ci, inv)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    cursor += 1
    if supply_data:
        for ci, inv in enumerate(INVESTOR_TYPES, 1):
            val = supply_data.get(inv, 0)
            display = int(round(val / 1_000_000)) if val else 0
            cell = ws.cell(cursor, ci, display)
            cell.border = THIN_BORDER
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
            if display > 0:
                cell.font = UP_FONT
            elif display < 0:
                cell.font = DOWN_FONT
    else:
        ws.cell(cursor, 1, "KRX 데이터 없음").font = NO_DATA_FONT
    cursor += 1

    # ── ■ 손익계산서 ──
    _write_section_header(ws, cursor, "■ 손익계산서", num_period_cols)
    cursor += 1
    is_start = cursor

    ws.cell(is_start, 1, "계정").font = HEADER_FONT
    ws.cell(is_start, 1).fill = HEADER_FILL
    ws.cell(is_start, 1).border = THIN_BORDER
    ws.cell(is_start, 1).alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 20

    for ci, period in enumerate(model.periods, start=2):
        cell = ws.cell(is_start, ci, period.display)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # IS row sequence
    row_defs: list[dict[str, Any]] = []
    for acct_def in IS_ACCOUNT_DEFS:
        row_defs.append({"type": "account", "def": acct_def})
        for mr in MARGIN_ROWS:
            if mr["after"] == acct_def["key"]:
                row_defs.append({"type": mr["type"], "label": mr["label"], "of": mr["of"]})

    rr = is_start + 1
    is_row_idx = 0
    for rd in row_defs:
        use_stripe = (is_row_idx % 2 == 1) and rd["type"] == "account"
        if rd["type"] == "account":
            adef = rd["def"]
            key = adef["key"]
            is_eps = key == "eps"
            lbl = ws.cell(rr, 1, adef["display"])
            lbl.font = LABEL_FONT
            lbl.border = THIN_BORDER
            if use_stripe:
                lbl.fill = STRIPE_FILL
            for ci, period in enumerate(model.periods, start=2):
                val = get_is_value(model, key, period)
                if is_eps:
                    display_val = f"{val:,}" if val else ""
                else:
                    display_val = to_eok(val) if val else ""
                cell = ws.cell(rr, ci, display_val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="right")
                if use_stripe:
                    cell.fill = STRIPE_FILL
                if not is_eps and isinstance(display_val, int):
                    cell.number_format = "#,##0"
            is_row_idx += 1
        elif rd["type"] == "margin":
            of_key = rd["of"]
            ws.cell(rr, 1, rd["label"]).font = PCT_FONT
            ws.cell(rr, 1).border = THIN_BORDER
            ws.cell(rr, 1).fill = LIGHT_FILL
            for ci, period in enumerate(model.periods, start=2):
                rev = get_is_value(model, "revenue", period)
                val = get_is_value(model, of_key, period)
                pct = safe_pct(val, rev)
                cell = ws.cell(rr, ci, round(pct, 1) if pct is not None else "")
                cell.font = PCT_FONT
                cell.border = THIN_BORDER
                cell.fill = LIGHT_FILL
                cell.alignment = Alignment(horizontal="right")
        elif rd["type"] == "yoy":
            of_key = rd["of"]
            ws.cell(rr, 1, rd["label"]).font = PCT_FONT
            ws.cell(rr, 1).border = THIN_BORDER
            ws.cell(rr, 1).fill = LIGHT_FILL
            for ci, period in enumerate(model.periods, start=2):
                yoy = get_yoy_value(model, of_key, period)
                cell = ws.cell(rr, ci, round(yoy, 1) if yoy is not None else "")
                cell.font = PCT_FONT
                cell.border = THIN_BORDER
                cell.fill = LIGHT_FILL
                cell.alignment = Alignment(horizontal="right")
        rr += 1
    cursor = rr

    # ── ■ 재무상태표 (요약) ──
    _write_section_header(ws, cursor, "■ 재무상태표 (요약)", num_period_cols)
    cursor += 1
    bs_start = cursor

    ws.cell(bs_start, 1, "계정").font = HEADER_FONT
    ws.cell(bs_start, 1).fill = HEADER_FILL
    ws.cell(bs_start, 1).border = THIN_BORDER
    ws.cell(bs_start, 1).alignment = Alignment(horizontal="center")
    for ci, period in enumerate(model.periods, start=2):
        cell = ws.cell(bs_start, ci, period.display)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    rr = bs_start + 1
    bs_row_idx = 0
    for acct_def in BS_ACCOUNT_DEFS:
        key = acct_def["key"]
        use_stripe = bs_row_idx % 2 == 1
        lbl = ws.cell(rr, 1, acct_def["display"])
        lbl.font = LABEL_FONT
        lbl.border = THIN_BORDER
        if use_stripe:
            lbl.fill = STRIPE_FILL
        for ci, period in enumerate(model.periods, start=2):
            val = get_bs_value(model, key, period)
            display_val = to_eok(val) if val else ""
            cell = ws.cell(rr, ci, display_val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right")
            if use_stripe:
                cell.fill = STRIPE_FILL
            if isinstance(display_val, int):
                cell.number_format = "#,##0"
        bs_row_idx += 1
        rr += 1

    # 유동비율(%)
    ws.cell(rr, 1, "  유동비율(%)").font = PCT_FONT
    ws.cell(rr, 1).border = THIN_BORDER
    ws.cell(rr, 1).fill = LIGHT_FILL
    for ci, period in enumerate(model.periods, start=2):
        ca = get_bs_value(model, "current_assets", period)
        cl = get_bs_value(model, "current_liabilities", period)
        pct = safe_pct(ca, cl)
        cell = ws.cell(rr, ci, round(pct, 1) if pct is not None else "")
        cell.font = PCT_FONT
        cell.border = THIN_BORDER
        cell.fill = LIGHT_FILL
        cell.alignment = Alignment(horizontal="right")
    rr += 1

    # 부채비율(%)
    ws.cell(rr, 1, "  부채비율(%)").font = PCT_FONT
    ws.cell(rr, 1).border = THIN_BORDER
    ws.cell(rr, 1).fill = LIGHT_FILL
    for ci, period in enumerate(model.periods, start=2):
        tl = get_bs_value(model, "total_liabilities", period)
        te = get_bs_value(model, "total_equity", period)
        pct = safe_pct(tl, te)
        cell = ws.cell(rr, ci, round(pct, 1) if pct is not None else "")
        cell.font = PCT_FONT
        cell.border = THIN_BORDER
        cell.fill = LIGHT_FILL
        cell.alignment = Alignment(horizontal="right")
    rr += 1
    cursor = rr

    # ── ■ 수익성 지표 ──
    _write_section_header(ws, cursor, "■ 수익성 지표", num_period_cols)
    cursor += 1
    prof_start = cursor

    ws.cell(prof_start, 1, "지표").font = HEADER_FONT
    ws.cell(prof_start, 1).fill = HEADER_FILL
    ws.cell(prof_start, 1).border = THIN_BORDER
    ws.cell(prof_start, 1).alignment = Alignment(horizontal="center")
    for ci, period in enumerate(model.periods, start=2):
        cell = ws.cell(prof_start, ci, period.display)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    rr = prof_start + 1
    prof_defs = [
        ("ROE(%)", "net_income", "total_equity"),
        ("ROA(%)", "net_income", "total_assets"),
        ("OPM(%)", "operating", "revenue"),
        ("NPM(%)", "net_income", "revenue"),
        ("매출원가율(%)", "cogs", "revenue"),
    ]
    for pi, (label, num_key, den_key) in enumerate(prof_defs):
        use_stripe = pi % 2 == 1
        lbl_cell = ws.cell(rr, 1, label)
        lbl_cell.font = LABEL_FONT
        lbl_cell.border = THIN_BORDER
        if use_stripe:
            lbl_cell.fill = STRIPE_FILL
        for ci, period in enumerate(model.periods, start=2):
            # 분자: IS 데이터, 분모: BS 또는 IS 데이터
            numerator = get_is_value(model, num_key, period)
            if den_key in ("total_equity", "total_assets"):
                denominator = get_bs_value(model, den_key, period)
            else:
                denominator = get_is_value(model, den_key, period)
            pct = safe_pct(numerator, denominator)
            cell = ws.cell(rr, ci, round(pct, 1) if pct is not None else "")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right")
            if use_stripe:
                cell.fill = STRIPE_FILL
        rr += 1
    cursor = rr

    # ── ■ 사업부 매출 (있을 때만) ──
    if segments:
        _write_section_header(ws, cursor, "■ 사업부 매출", 4)
        cursor += 1
        seg_headers = ["사업부명", "매출액(억)", "비중(%)"]
        for ci, h in enumerate(seg_headers, 1):
            cell = ws.cell(cursor, ci, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        cursor += 1
        for si, seg in enumerate(segments):
            use_stripe = si % 2 == 1
            ws.cell(cursor, 1, seg["name"]).border = THIN_BORDER
            if use_stripe:
                ws.cell(cursor, 1).fill = STRIPE_FILL
            c2 = ws.cell(cursor, 2, to_eok(seg["revenue"]))
            c2.border = THIN_BORDER
            c2.number_format = "#,##0"
            c2.alignment = Alignment(horizontal="right")
            if use_stripe:
                c2.fill = STRIPE_FILL
            c3 = ws.cell(cursor, 3, round(seg["pct"], 1))
            c3.border = THIN_BORDER
            c3.alignment = Alignment(horizontal="right")
            if use_stripe:
                c3.fill = STRIPE_FILL
            cursor += 1

    # ── ■ 최근 공시 ──
    _write_section_header(ws, cursor, "■ 최근 공시", 4)
    cursor += 1
    disc_headers = ["일자", "공시명", "DART 링크"]
    for ci, h in enumerate(disc_headers, 1):
        cell = ws.cell(cursor, ci, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    cursor += 1
    if disclosures:
        for i, disc in enumerate(disclosures):
            use_stripe = i % 2 == 1
            # 일자
            c1 = ws.cell(cursor, 1, disc.get("rcept_dt", ""))
            c1.border = THIN_BORDER
            if use_stripe:
                c1.fill = STRIPE_FILL
            # 공시명
            c2 = ws.cell(cursor, 2, disc.get("report_nm", ""))
            c2.border = THIN_BORDER
            if use_stripe:
                c2.fill = STRIPE_FILL
            # 링크
            rcept_no = disc.get("rcept_no", "")
            if rcept_no:
                link_url = f"{DART_REPORT_URL}{rcept_no}"
                c3 = ws.cell(cursor, 3, "보기")
                c3.hyperlink = link_url
                c3.font = LINK_FONT
            else:
                c3 = ws.cell(cursor, 3, "")
            c3.border = THIN_BORDER
            c3.alignment = Alignment(horizontal="center")
            if use_stripe:
                c3.fill = STRIPE_FILL
            cursor += 1
        max_len = max((len(d.get("report_nm", "")) for d in disclosures), default=20)
        ws.column_dimensions["B"].width = min(max(max_len * 1.3, 20), 60)
    else:
        ws.cell(cursor, 1, "공시 데이터 없음").font = NO_DATA_FONT

    # 열 너비 설정
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = f"B{is_start + 1}"
    ws.sheet_properties.tabColor = "2F5496"


def write_sheet_summary(
    ws: Any,
    target_date: date,
    models: list[FinancialModel],
    watchlist: list[dict[str, str]],
    price_by_code: dict[str, dict] | None = None,
) -> None:
    ws.cell(1, 1, f"{title_date_text(target_date)} 관심종목 요약").font = TITLE_FONT

    headers = [
        "종목명", "최근분기", "매출액(억)", "영업이익(억)", "OPM(%)",
        "당기순이익(억)", "NPM(%)", "매출 YoY(%)", "영업이익 YoY(%)",
        "PER(배)", "PBR(배)", "ROE(%)",
    ]
    set_header_row(ws, 3, headers)

    pbc = price_by_code or {}

    for i, (entry, model) in enumerate(zip(watchlist, models)):
        rr = i + 4
        stock_code = entry.get("stock_code", "")
        recent = None
        for p in reversed(model.periods):
            if p.quarter != "FY" and get_is_value(model, "revenue", p):
                recent = p
                break
        if not recent:
            ws.cell(rr, 1, entry.get("name", "")).border = THIN_BORDER
            ws.cell(rr, 2, "데이터 없음").border = THIN_BORDER
            continue

        rev = get_is_value(model, "revenue", recent)
        op = get_is_value(model, "operating", recent)
        ni = get_is_value(model, "net_income", recent)
        opm = safe_pct(op, rev)
        npm = safe_pct(ni, rev)
        rev_yoy = get_yoy_value(model, "revenue", recent)
        op_yoy = get_yoy_value(model, "operating", recent)

        # 밸류에이션
        vals = calc_valuations(model, pbc.get(stock_code))
        per_val = vals.get("PER")
        pbr_val = vals.get("PBR")

        # ROE (최근 분기)
        equity = get_bs_value(model, "total_equity", recent)
        roe = safe_pct(ni, equity) if equity else None

        ws.cell(rr, 1, entry.get("name", "")).border = THIN_BORDER
        ws.cell(rr, 2, recent.display).border = THIN_BORDER
        ws.cell(rr, 2).alignment = Alignment(horizontal="center")
        ws.cell(rr, 3, to_eok(rev)).border = THIN_BORDER
        ws.cell(rr, 3).number_format = "#,##0"
        ws.cell(rr, 4, to_eok(op)).border = THIN_BORDER
        ws.cell(rr, 4).number_format = "#,##0"
        ws.cell(rr, 5, round(opm, 1) if opm is not None else "").border = THIN_BORDER
        ws.cell(rr, 6, to_eok(ni)).border = THIN_BORDER
        ws.cell(rr, 6).number_format = "#,##0"
        ws.cell(rr, 7, round(npm, 1) if npm is not None else "").border = THIN_BORDER
        ws.cell(rr, 8, round(rev_yoy, 1) if rev_yoy is not None else "").border = THIN_BORDER
        ws.cell(rr, 9, round(op_yoy, 1) if op_yoy is not None else "").border = THIN_BORDER
        ws.cell(rr, 10, round(per_val, 1) if per_val is not None else "").border = THIN_BORDER
        ws.cell(rr, 11, round(pbr_val, 1) if pbr_val is not None else "").border = THIN_BORDER
        ws.cell(rr, 12, round(roe, 1) if roe is not None else "").border = THIN_BORDER

    widths = [14, 10, 14, 14, 10, 14, 10, 12, 14, 10, 10, 10]
    for c in range(1, len(widths) + 1):
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    ws.freeze_panes = "A4"


def write_sheet_price(
    ws: Any,
    target_date: date,
    price_rows: list[dict],
    watchlist: list[dict[str, str]],
) -> None:
    ws.cell(1, 1, f"{title_date_text(target_date)} 관심종목 시세").font = TITLE_FONT

    if not price_rows:
        ws.cell(3, 1, "KRX 접속 불가 또는 데이터 없음")
        return

    headers = ["종목코드", "종목명", "종가", "등락(원)", "등락률(%)", "거래량", "거래대금(억)", "시가총액(억)"]
    set_header_row(ws, 3, headers)
    widths = [12, 16, 14, 12, 10, 14, 14, 16]

    code_to_name = {w["stock_code"]: w["name"] for w in watchlist}

    for i, row in enumerate(price_rows, start=4):
        code = str(row.get("ISU_SRT_CD", row.get("종목코드", ""))).strip()
        ws.cell(i, 1, code).border = THIN_BORDER
        ws.cell(i, 2, code_to_name.get(code, row.get("ISU_ABBRV", ""))).border = THIN_BORDER
        ws.cell(i, 3, parse_amount(row.get("TDD_CLSPRC", row.get("종가", "0")))).border = THIN_BORDER
        ws.cell(i, 3).number_format = "#,##0"
        ws.cell(i, 4, parse_amount(row.get("CMPPREVDD_PRC", row.get("대비", "0")))).border = THIN_BORDER
        ws.cell(i, 4).number_format = "#,##0"
        pct_str = str(row.get("FLUC_RT", row.get("등락률", "0"))).replace(",", "").replace("%", "")
        try:
            pct_val = float(pct_str)
        except ValueError:
            pct_val = 0.0
        ws.cell(i, 5, round(pct_val, 2)).border = THIN_BORDER
        ws.cell(i, 6, parse_amount(row.get("ACC_TRDVOL", row.get("거래량", "0")))).border = THIN_BORDER
        ws.cell(i, 6).number_format = "#,##0"
        ws.cell(i, 7, to_eok(parse_amount(row.get("ACC_TRDVAL", row.get("거래대금", "0"))))).border = THIN_BORDER
        ws.cell(i, 7).number_format = "#,##0"
        ws.cell(i, 8, to_eok(parse_amount(row.get("MKTCAP", row.get("시가총액", "0"))))).border = THIN_BORDER
        ws.cell(i, 8).number_format = "#,##0"

    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"


def write_sheet_supply(
    ws: Any,
    target_date: date,
    supply_by_type: dict[str, list[dict]],
    watchlist: list[dict[str, str]],
) -> None:
    ws.cell(1, 1, f"{title_date_text(target_date)} 관심종목 수급").font = TITLE_FONT
    ws.cell(2, 1, "단위: 백만원 (순매수거래대금)").font = Font(size=9, color="888888")

    if not supply_by_type:
        ws.cell(3, 1, "KRX 접속 불가 또는 데이터 없음")
        return

    headers = ["종목코드", "종목명"] + INVESTOR_TYPES
    set_header_row(ws, 3, headers)

    code_to_name = {w["stock_code"]: w["name"] for w in watchlist}
    all_codes = [w["stock_code"] for w in watchlist]

    supply_map: dict[str, dict[str, int]] = {c: {} for c in all_codes}
    for inv_name, rows in supply_by_type.items():
        for row in rows:
            code = str(row.get("ISU_SRT_CD", row.get("종목코드", ""))).strip()
            if code in supply_map:
                net_str = row.get("NETBID_TRDVAL", row.get("순매수거래대금", "0"))
                supply_map[code][inv_name] = parse_amount(net_str)

    for i, code in enumerate(all_codes, start=4):
        ws.cell(i, 1, code).border = THIN_BORDER
        ws.cell(i, 2, code_to_name.get(code, "")).border = THIN_BORDER
        for j, inv in enumerate(INVESTOR_TYPES, start=3):
            val = supply_map.get(code, {}).get(inv, 0)
            cell = ws.cell(i, j, int(round(val / 1_000_000)) if val else 0)
            cell.border = THIN_BORDER
            cell.number_format = "#,##0"

    widths = [12, 16] + [14] * len(INVESTOR_TYPES)
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"


def write_watchlist_workbook(
    out_path: Path,
    target_date: date,
    models: list[FinancialModel],
    watchlist: list[dict[str, str]],
    price_rows: list[dict],
    supply_by_type: dict[str, list[dict]],
    disclosures_by_corp: dict[str, list[dict]] | None = None,
    segments_by_corp: dict[str, list[dict] | None] | None = None,
) -> None:
    wb = Workbook()

    # Build per-stock lookup maps
    price_by_code: dict[str, dict] = {}
    for row in price_rows:
        code = str(row.get("ISU_SRT_CD", row.get("종목코드", ""))).strip()
        price_by_code[code] = row

    supply_by_code: dict[str, dict[str, int]] = {}
    for inv_name, rows in supply_by_type.items():
        for row in rows:
            code = str(row.get("ISU_SRT_CD", row.get("종목코드", ""))).strip()
            if code not in supply_by_code:
                supply_by_code[code] = {}
            net_str = row.get("NETBID_TRDVAL", row.get("순매수거래대금", "0"))
            supply_by_code[code][inv_name] = parse_amount(net_str)

    # 요약 시트
    ws_summary = wb.active
    ws_summary.title = "요약"
    write_sheet_summary(ws_summary, target_date, models, watchlist, price_by_code)

    # 종목별 시트
    seg_map = segments_by_corp or {}
    for entry, model in zip(watchlist, models):
        name = entry.get("name", "")
        stock_code = entry.get("stock_code", "")
        corp_code = entry.get("corp_code", "")
        sheet_name = name[:31]
        ws = wb.create_sheet(title=sheet_name)
        write_sheet_equity(
            ws, model, target_date,
            price_row=price_by_code.get(stock_code),
            supply_data=supply_by_code.get(stock_code),
            disclosures=(disclosures_by_corp or {}).get(corp_code),
            segments=seg_map.get(corp_code),
        )

    ws_price = wb.create_sheet(title="시세")
    write_sheet_price(ws_price, target_date, price_rows, watchlist)

    ws_supply = wb.create_sheet(title="수급")
    write_sheet_supply(ws_supply, target_date, supply_by_type, watchlist)

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    args = parse_args()
    config_path = Path(args.config).resolve()
    ensure_file(config_path, "설정")
    cfg = load_json(config_path)

    dart_cfg = cfg.get("dart", {})
    api_key = dart_cfg.get("api_key", "")
    if not api_key:
        raise ValueError("config에 dart.api_key가 설정되어 있지 않습니다.")

    watchlist = cfg.get("watchlist", [])
    if not watchlist:
        print("[경고] watchlist가 비어있습니다.")
        return 0

    target_date = parse_target_date(args.date)
    yyyymmdd = to_yyyymmdd(target_date)
    yymmdd = to_yymmdd(target_date)

    wo = cfg.get("watchlist_output", {})
    financial_years = wo.get("financial_years", [target_date.year - 1, target_date.year])
    fs_div = wo.get("fs_div", "CFS")

    client = DartFinancialClient(
        api_key=api_key,
        timeout_sec=int(dart_cfg.get("timeout_sec", 30)),
        delay=float(dart_cfg.get("request_delay_sec", 0.5)),
    )

    print(f"[시작] {title_date_text(target_date)} 관심종목 Equity Research Model 분석")

    # 1) 재무 모델 수집 (IS + BS)
    models: list[FinancialModel] = []
    for entry in watchlist:
        name = entry.get("name", "")
        corp_code = entry.get("corp_code", "")
        stock_code = entry.get("stock_code", "")
        print(f"[재무] {name} IS+BS 수집 중...")
        model = build_financial_model(
            client, corp_code, name, stock_code, financial_years, fs_div,
        )
        models.append(model)
        rev_fy = model.is_data.get("revenue", {}).get(f"{financial_years[-1]}_FY", 0)
        equity_fy = model.bs_data.get("total_equity", {}).get(f"{financial_years[-1]}_FY", 0)
        if rev_fy:
            print(f"  → FY{financial_years[-1]} 매출액: {to_eok(rev_fy):,}억")
        else:
            print(f"  → FY{financial_years[-1]} 데이터 없음 (미공시)")
        if equity_fy:
            print(f"  → FY{financial_years[-1]} 자본총계: {to_eok(equity_fy):,}억")

    # 2) 사업부 매출 수집 (DART API, 실패시 생략)
    segments_by_corp: dict[str, list[dict] | None] = {}
    for entry in watchlist:
        name = entry.get("name", "")
        corp_code = entry.get("corp_code", "")
        print(f"[사업부] {name} 사업부 매출 조회 중...")
        segs = fetch_segment_data(client, corp_code, financial_years[-1], fs_div=fs_div)
        segments_by_corp[corp_code] = segs
        if segs:
            print(f"  → {len(segs)}개 사업부 발견")
        else:
            print(f"  → 사업부 데이터 없음 (생략)")

    # 3) KRX 데이터 (선택적)
    krx_config_path = Path(args.krx_config).resolve()
    watchlist_codes = {w["stock_code"] for w in watchlist}
    price_rows, supply_by_type = try_fetch_krx_data(krx_config_path, yyyymmdd, watchlist_codes)

    # 4) 종목별 최근 공시 수집
    disclosures_by_corp: dict[str, list[dict]] = {}
    bgn_de = to_yyyymmdd(target_date - timedelta(days=30))
    end_de = yyyymmdd
    for entry in watchlist:
        name = entry.get("name", "")
        corp_code = entry.get("corp_code", "")
        print(f"[공시] {name} 최근 공시 조회 중...")
        discs = client.fetch_disclosures(corp_code, bgn_de, end_de, page_count=10)
        disclosures_by_corp[corp_code] = discs
        if discs:
            print(f"  → {len(discs)}건")
        else:
            print(f"  → 없음")

    # 5) 엑셀 출력
    output_dir = Path(cfg.get("output_dir", "시장 정리")).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{yymmdd}_관심종목.xlsx"

    write_watchlist_workbook(
        out_path=out_path,
        target_date=target_date,
        models=models,
        watchlist=watchlist,
        price_rows=price_rows,
        supply_by_type=supply_by_type,
        disclosures_by_corp=disclosures_by_corp,
        segments_by_corp=segments_by_corp,
    )

    print(f"[완료] {out_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"[오류] {exc}", file=sys.stderr)
        raise
