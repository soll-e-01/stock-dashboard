import argparse
import json
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]

CORP_CLS_ORDER = {"Y": 1, "K": 2, "N": 3, "E": 4}
CORP_CLS_DISPLAY = {"Y": "유가", "K": "코스닥", "N": "코넥스", "E": "기타"}

PBLNTF_TY_DISPLAY = {
    "A": "정기공시",
    "B": "주요사항",
    "C": "발행공시",
    "D": "지분공시",
    "E": "기타공시",
    "F": "외부감사",
    "G": "펀드공시",
    "H": "자산유동화",
    "I": "거래소공시",
    "J": "공정위공시",
}

SHEET_TYPE_MAP: dict[str, list[str]] = {
    "주요사항": ["B"],
    "지분공시": ["D"],
    "공정공시": ["I", "E"],
}

TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
FONT_UP = Font(color="CC0000", bold=True)      # 상승: 빨간색
FONT_DOWN = Font(color="0000CC", bold=True)     # 하락: 파란색
FONT_FLAT = Font(color="666666")                # 보합: 회색


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="DART 공시 일일 데이터 수집 후 엑셀 생성")
    p.add_argument("--config", default="config.dart.json", help="설정 파일 경로")
    p.add_argument("--date", default="", help="대상 일자 YYYYMMDD (미입력 시 오늘)")
    p.add_argument(
        "--lookup",
        default="",
        help="종목코드로 corp_code 조회 (예: --lookup 005930)",
    )
    return p.parse_args()


def parse_target_date(raw: str) -> date:
    if raw:
        return datetime.strptime(raw, "%Y%m%d").date()
    return date.today()


def to_yyyymmdd(d: date) -> str:
    return d.strftime("%Y%m%d")


def to_yymmdd(d: date) -> str:
    return d.strftime("%y%m%d")


def title_date_text(d: date) -> str:
    return f"{d:%Y/%m/%d}({WEEKDAY_KR[d.weekday()]})"


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def ensure_file(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{label} 파일이 없습니다: {path}")


def as_text_code(v: Any) -> str:
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if s.isdigit() and len(s) <= 6:
        return s.zfill(6)
    return s


def load_price_data(
    output_dir: Path, yymmdd: str,
) -> tuple[dict[str, int], dict[str, float]]:
    """시세_DATA 엑셀에서 종목코드별 시가총액(원)과 등락률(%) 로드."""
    path = output_dir / f"{yymmdd}_시세_DATA.xlsx"
    if not path.exists():
        candidates = sorted(output_dir.glob("*_시세_DATA.xlsx"), reverse=True)
        if not candidates:
            return {}, {}
        path = candidates[0]
        print(f"[시세] 최근 시세 파일 사용: {path.name}")
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if "DATA" not in wb.sheetnames:
            wb.close()
            return {}, {}
        ws = wb["DATA"]
        cap_result: dict[str, int] = {}
        pct_result: dict[str, float] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            code = str(row[0]).strip()
            if code.isdigit():
                code = code.zfill(6)
            # col 12 (idx 12) = 시가총액
            market_cap = row[12] if len(row) > 12 else None
            if market_cap:
                try:
                    cap_result[code] = int(market_cap)
                except (ValueError, TypeError):
                    pass
            # col 6 (idx 6) = 등락률
            pct = row[6] if len(row) > 6 else None
            if pct is not None:
                try:
                    pct_result[code] = float(pct)
                except (ValueError, TypeError):
                    pass
        wb.close()
        return cap_result, pct_result
    except Exception as exc:
        print(f"[시세] 시세 파일 읽기 실패: {exc}")
        return {}, {}


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class DisclosureRow:
    corp_code: str
    corp_name: str
    stock_code: str
    corp_cls: str
    report_nm: str
    rcept_no: str
    flr_nm: str
    rcept_dt: str
    rm: str
    pblntf_ty: str
    subcategory: str = ""
    subcategory_priority: int = 99


@dataclass
class CategorySummary:
    pblntf_ty: str
    display_name: str
    total_count: int = 0
    subcategory_counts: dict[str, int] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# DART Client
# ---------------------------------------------------------------------------

class DartClient:
    def __init__(
        self,
        api_key: str,
        base_url: str = "https://opendart.fss.or.kr/api",
        timeout_sec: int = 30,
        page_size: int = 100,
        request_delay: float = 0.5,
    ):
        self.api_key = api_key
        self.base_url = base_url.rstrip("/")
        self.timeout_sec = timeout_sec
        self.page_size = page_size
        self.request_delay = request_delay
        self.session = requests.Session()

    def fetch_disclosures(
        self,
        bgn_de: str,
        end_de: str,
        pblntf_ty: str = "",
        corp_code: str = "",
    ) -> list[dict[str, str]]:
        all_rows: list[dict[str, str]] = []
        page_no = 1

        while True:
            params: dict[str, str] = {
                "crtfc_key": self.api_key,
                "bgn_de": bgn_de,
                "end_de": end_de,
                "page_no": str(page_no),
                "page_count": str(self.page_size),
                "sort": "date",
                "sort_mth": "desc",
            }
            if pblntf_ty:
                params["pblntf_ty"] = pblntf_ty
            if corp_code:
                params["corp_code"] = corp_code

            resp = self.session.get(
                f"{self.base_url}/list.json",
                params=params,
                timeout=self.timeout_sec,
            )
            resp.raise_for_status()
            body = resp.json()

            status = body.get("status", "")
            if status == "013":
                break
            if status != "000":
                msg = body.get("message", "알 수 없는 오류")
                raise RuntimeError(f"DART API 오류 [{status}]: {msg}")

            rows = body.get("list", [])
            all_rows.extend(rows)

            total_page = int(body.get("total_page", 1))
            if page_no >= total_page:
                break

            page_no += 1
            time.sleep(self.request_delay)

        return all_rows

    def fetch_all_types(
        self,
        bgn_de: str,
        end_de: str,
        pblntf_types: list[str],
    ) -> list[dict[str, str]]:
        all_rows: list[dict[str, str]] = []
        for ty in pblntf_types:
            rows = self.fetch_disclosures(bgn_de=bgn_de, end_de=end_de, pblntf_ty=ty)
            for row in rows:
                row["_pblntf_ty"] = ty
            all_rows.extend(rows)
            time.sleep(self.request_delay)
        return all_rows

    def fetch_watchlist(
        self,
        bgn_de: str,
        end_de: str,
        watchlist: list[dict[str, str]],
    ) -> list[dict[str, str]]:
        def _normalize_corp_code(v: object) -> str:
            s = str(v or "").strip()
            if s.startswith("'"):
                s = s[1:]
            # Remove whitespace / hidden chars / non-digits from sheet exports.
            s = re.sub(r"\D", "", s)
            if not s:
                return ""
            if len(s) > 8:
                return ""
            return s.zfill(8)

        all_rows: list[dict[str, str]] = []
        for entry in watchlist:
            raw_corp_code = entry.get("corp_code", "")
            corp_code = _normalize_corp_code(raw_corp_code)
            name = entry.get("name", "")
            if not corp_code:
                print(f"[경고] {name}: corp_code 형식 오류로 건너뜁니다. raw={raw_corp_code!r}")
                continue
            try:
                rows = self.fetch_disclosures(
                    bgn_de=bgn_de, end_de=end_de, corp_code=corp_code,
                )
            except Exception as e:
                print(f"[경고] {name}({corp_code}) 공시 조회 실패: {e}")
                continue
            all_rows.extend(rows)
            time.sleep(self.request_delay)
        return all_rows

    def lookup_corp_code(self, stock_code: str) -> None:
        import io
        import zipfile
        from xml.etree import ElementTree as ET

        print(f"[조회] DART corpCode.xml 다운로드 중...")
        resp = self.session.get(
            f"{self.base_url}/corpCode.xml",
            params={"crtfc_key": self.api_key},
            timeout=60,
        )
        resp.raise_for_status()

        with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
            xml_name = zf.namelist()[0]
            root = ET.fromstring(zf.read(xml_name))

        target = stock_code.strip().zfill(6)
        found = False
        for item in root.iter("list"):
            sc = (item.findtext("stock_code") or "").strip()
            if sc == target:
                cc = (item.findtext("corp_code") or "").strip()
                cn = (item.findtext("corp_name") or "").strip()
                print(f"  종목코드: {sc}")
                print(f"  고유번호(corp_code): {cc}")
                print(f"  회사명: {cn}")
                print(f'  config 항목: {{"stock_code": "{sc}", "corp_code": "{cc}", "name": "{cn}"}}')
                found = True
                break
        if not found:
            print(f"[오류] 종목코드 {target}에 해당하는 회사를 찾지 못했습니다.")


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

def classify_disclosure(
    report_nm: str,
    pblntf_ty: str,
    classification: dict[str, Any],
) -> tuple[str, int]:
    categories = classification.get(pblntf_ty, [])
    for entry in sorted(categories, key=lambda x: x.get("priority", 99)):
        if entry.get("is_fallback"):
            continue
        for kw in entry.get("keywords", []):
            if re.search(kw, report_nm):
                return entry["subcategory"], entry.get("priority", 99)
    for entry in categories:
        if entry.get("is_fallback"):
            return entry["subcategory"], entry.get("priority", 99)
    return PBLNTF_TY_DISPLAY.get(pblntf_ty, "기타"), 99


def parse_disclosure_rows(
    raw_rows: list[dict[str, str]],
    classification: dict[str, Any],
) -> list[DisclosureRow]:
    seen: set[str] = set()
    out: list[DisclosureRow] = []
    for row in raw_rows:
        rcept_no = row.get("rcept_no", "")
        if rcept_no in seen:
            continue
        seen.add(rcept_no)

        pblntf_ty = row.get("_pblntf_ty", row.get("pblntf_ty", ""))
        report_nm = row.get("report_nm", "")
        sub, pri = classify_disclosure(report_nm, pblntf_ty, classification)

        out.append(DisclosureRow(
            corp_code=row.get("corp_code", ""),
            corp_name=row.get("corp_name", ""),
            stock_code=as_text_code(row.get("stock_code", "")),
            corp_cls=row.get("corp_cls", ""),
            report_nm=report_nm,
            rcept_no=rcept_no,
            flr_nm=row.get("flr_nm", ""),
            rcept_dt=row.get("rcept_dt", ""),
            rm=row.get("rm", ""),
            pblntf_ty=pblntf_ty,
            subcategory=sub,
            subcategory_priority=pri,
        ))
    return out


def build_summaries(
    rows: list[DisclosureRow],
    sheet_type_map: dict[str, list[str]],
) -> list[CategorySummary]:
    summaries: list[CategorySummary] = []
    for sheet_name, types in sheet_type_map.items():
        filtered = [r for r in rows if r.pblntf_ty in types]
        sub_counts: dict[str, int] = {}
        for r in filtered:
            sub_counts[r.subcategory] = sub_counts.get(r.subcategory, 0) + 1
        summaries.append(CategorySummary(
            pblntf_ty="/".join(types),
            display_name=sheet_name,
            total_count=len(filtered),
            subcategory_counts=sub_counts,
        ))
    # 정기공시
    periodic = [r for r in rows if r.pblntf_ty == "A"]
    sub_counts_a: dict[str, int] = {}
    for r in periodic:
        sub_counts_a[r.subcategory] = sub_counts_a.get(r.subcategory, 0) + 1
    summaries.append(CategorySummary(
        pblntf_ty="A",
        display_name="정기공시",
        total_count=len(periodic),
        subcategory_counts=sub_counts_a,
    ))
    return summaries


def group_by_sheet(
    rows: list[DisclosureRow],
    sheet_type_map: dict[str, list[str]],
) -> dict[str, list[DisclosureRow]]:
    groups: dict[str, list[DisclosureRow]] = {}
    for sheet_name, types in sheet_type_map.items():
        filtered = [r for r in rows if r.pblntf_ty in types]
        filtered.sort(key=lambda r: (
            r.subcategory_priority,
            CORP_CLS_ORDER.get(r.corp_cls, 9),
        ))
        groups[sheet_name] = filtered
    return groups


# ---------------------------------------------------------------------------
# Excel Writers
# ---------------------------------------------------------------------------

def set_header_row(ws: Any, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def set_col_widths(ws: Any, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)
        ws.column_dimensions[col_letter].width = w


def write_pct_cell(ws: Any, row: int, col: int, pct: float | None) -> None:
    """등락률 셀을 색상 코딩하여 기록."""
    cell = ws.cell(row, col)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="right")
    if pct is None:
        return
    cell.value = round(pct, 2)
    cell.number_format = "0.00"
    if pct > 0:
        cell.font = FONT_UP
    elif pct < 0:
        cell.font = FONT_DOWN
    else:
        cell.font = FONT_FLAT


def write_sheet_summary(
    ws: Any,
    target_date: date,
    summaries: list[CategorySummary],
    watchlist_rows: list[DisclosureRow],
    pct_by_code: dict[str, float] | None = None,
) -> None:
    ws.cell(1, 1, f"{title_date_text(target_date)} DART 공시 일일 요약").font = TITLE_FONT

    set_header_row(ws, 3, ["구분", "건수", "세부 분류 현황"])
    set_col_widths(ws, [14, 8, 80])

    row = 4
    total = 0
    for s in summaries:
        ws.cell(row, 1, s.display_name).border = THIN_BORDER
        ws.cell(row, 2, s.total_count).border = THIN_BORDER
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
        sorted_subs = sorted(s.subcategory_counts.items(), key=lambda x: -x[1])
        detail = ", ".join(f"{k}({v})" for k, v in sorted_subs[:8])
        ws.cell(row, 3, detail).border = THIN_BORDER
        total += s.total_count
        row += 1

    ws.cell(row, 1, "합계").font = HEADER_FONT
    ws.cell(row, 1).border = THIN_BORDER
    ws.cell(row, 2, total).font = HEADER_FONT
    ws.cell(row, 2).border = THIN_BORDER
    ws.cell(row, 2).alignment = Alignment(horizontal="center")
    ws.cell(row, 3).border = THIN_BORDER

    row += 2
    ws.cell(row, 1, "관심종목 공시 현황").font = Font(bold=True, size=12)
    row += 1
    pct = pct_by_code or {}
    if watchlist_rows:
        headers = ["종목코드", "종목명", "등락률(%)", "공시유형", "세부분류", "보고서명", "비고"]
        set_header_row(ws, row, headers)
        set_col_widths(ws, [14, 14, 10, 12, 14, 60, 8])
        row += 1
        for r in watchlist_rows:
            ws.cell(row, 1, r.stock_code).border = THIN_BORDER
            ws.cell(row, 2, r.corp_name).border = THIN_BORDER
            stock_pct = pct.get(r.stock_code)
            write_pct_cell(ws, row, 3, stock_pct)
            ws.cell(row, 4, PBLNTF_TY_DISPLAY.get(r.pblntf_ty, r.pblntf_ty)).border = THIN_BORDER
            ws.cell(row, 5, r.subcategory).border = THIN_BORDER
            ws.cell(row, 6, r.report_nm).border = THIN_BORDER
            ws.cell(row, 7, r.rm).border = THIN_BORDER
            row += 1
    else:
        ws.cell(row, 1, "해당일 관심종목 공시 없음")

    ws.freeze_panes = "A4"


DISCLOSURE_HEADERS = ["No.", "시장구분", "종목코드", "종목명", "시총(억)", "등락률(%)", "세부분류", "보고서명", "제출인", "비고"]
DISCLOSURE_WIDTHS = [6, 10, 12, 18, 14, 10, 16, 55, 18, 8]


def write_sheet_disclosure(
    ws: Any,
    target_date: date,
    sheet_title: str,
    rows: list[DisclosureRow],
    market_cap_by_code: dict[str, int] | None = None,
    pct_by_code: dict[str, float] | None = None,
) -> None:
    ws.cell(1, 1, f"{title_date_text(target_date)} {sheet_title}").font = TITLE_FONT

    set_header_row(ws, 3, DISCLOSURE_HEADERS)
    set_col_widths(ws, DISCLOSURE_WIDTHS)

    if not rows:
        ws.cell(4, 1, "해당일 공시 없음")
        ws.freeze_panes = "A4"
        return

    mcap = market_cap_by_code or {}
    pct = pct_by_code or {}
    for i, r in enumerate(rows, start=1):
        rr = i + 3
        ws.cell(rr, 1, i).border = THIN_BORDER
        ws.cell(rr, 1).alignment = Alignment(horizontal="center")
        ws.cell(rr, 2, CORP_CLS_DISPLAY.get(r.corp_cls, r.corp_cls)).border = THIN_BORDER
        ws.cell(rr, 3, r.stock_code).border = THIN_BORDER
        ws.cell(rr, 4, r.corp_name).border = THIN_BORDER
        cap = mcap.get(r.stock_code, 0)
        c5 = ws.cell(rr, 5, int(round(cap / 1_0000_0000)) if cap else None)
        c5.border = THIN_BORDER
        if cap:
            c5.number_format = "#,##0"
            c5.alignment = Alignment(horizontal="right")
        stock_pct = pct.get(r.stock_code)
        write_pct_cell(ws, rr, 6, stock_pct)
        ws.cell(rr, 7, r.subcategory).border = THIN_BORDER
        ws.cell(rr, 8, r.report_nm).border = THIN_BORDER
        ws.cell(rr, 9, r.flr_nm).border = THIN_BORDER
        ws.cell(rr, 10, r.rm).border = THIN_BORDER

    last_row = len(rows) + 3
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:J{last_row}"


WATCHLIST_HEADERS = ["No.", "종목코드", "종목명", "시총(억)", "등락률(%)", "공시유형", "세부분류", "보고서명", "제출인", "접수일", "비고"]
WATCHLIST_WIDTHS = [6, 12, 18, 14, 10, 12, 16, 55, 18, 12, 8]


def write_sheet_watchlist(
    ws: Any,
    target_date: date,
    rows: list[DisclosureRow],
    market_cap_by_code: dict[str, int] | None = None,
    pct_by_code: dict[str, float] | None = None,
) -> None:
    ws.cell(1, 1, f"{title_date_text(target_date)} 관심종목 공시 상세").font = TITLE_FONT

    set_header_row(ws, 3, WATCHLIST_HEADERS)
    set_col_widths(ws, WATCHLIST_WIDTHS)

    if not rows:
        ws.cell(4, 1, "해당일 관심종목 공시 없음")
        ws.freeze_panes = "A4"
        return

    sorted_rows = sorted(rows, key=lambda r: (
        r.stock_code,
        r.subcategory_priority,
    ))

    mcap = market_cap_by_code or {}
    pct = pct_by_code or {}
    for i, r in enumerate(sorted_rows, start=1):
        rr = i + 3
        ws.cell(rr, 1, i).border = THIN_BORDER
        ws.cell(rr, 1).alignment = Alignment(horizontal="center")
        ws.cell(rr, 2, r.stock_code).border = THIN_BORDER
        ws.cell(rr, 3, r.corp_name).border = THIN_BORDER
        cap = mcap.get(r.stock_code, 0)
        c4 = ws.cell(rr, 4, int(round(cap / 1_0000_0000)) if cap else None)
        c4.border = THIN_BORDER
        if cap:
            c4.number_format = "#,##0"
            c4.alignment = Alignment(horizontal="right")
        stock_pct = pct.get(r.stock_code)
        write_pct_cell(ws, rr, 5, stock_pct)
        ws.cell(rr, 6, PBLNTF_TY_DISPLAY.get(r.pblntf_ty, r.pblntf_ty)).border = THIN_BORDER
        ws.cell(rr, 7, r.subcategory).border = THIN_BORDER
        ws.cell(rr, 8, r.report_nm).border = THIN_BORDER
        ws.cell(rr, 9, r.flr_nm).border = THIN_BORDER
        ws.cell(rr, 10, r.rcept_dt).border = THIN_BORDER
        ws.cell(rr, 11, r.rm).border = THIN_BORDER

    last_row = len(sorted_rows) + 3
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K{last_row}"


def write_dart_workbook(
    out_path: Path,
    target_date: date,
    summaries: list[CategorySummary],
    by_sheet: dict[str, list[DisclosureRow]],
    watchlist: list[DisclosureRow],
    market_cap_by_code: dict[str, int] | None = None,
    pct_by_code: dict[str, float] | None = None,
) -> None:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "요약"
    write_sheet_summary(ws_summary, target_date, summaries, watchlist, pct_by_code)

    mcap = market_cap_by_code or {}
    pct = pct_by_code or {}
    for sheet_name in ["주요사항", "지분공시", "공정공시"]:
        ws = wb.create_sheet(title=sheet_name)
        write_sheet_disclosure(ws, target_date, sheet_name, by_sheet.get(sheet_name, []), mcap, pct)

    ws_watch = wb.create_sheet(title="관심종목")
    write_sheet_watchlist(ws_watch, target_date, watchlist, mcap, pct)

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    args = parse_args()
    config_path = Path(args.config).resolve()
    ensure_file(config_path, "설정")
    cfg = load_json(config_path)

    dart_cfg = cfg.get("dart", {})
    api_key = dart_cfg.get("api_key", "")
    if not api_key:
        raise ValueError("config에 dart.api_key가 설정되어 있지 않습니다.")

    client = DartClient(
        api_key=api_key,
        base_url=dart_cfg.get("base_url", "https://opendart.fss.or.kr/api"),
        timeout_sec=int(dart_cfg.get("timeout_sec", 30)),
        page_size=int(dart_cfg.get("page_size", 100)),
        request_delay=float(dart_cfg.get("request_delay_sec", 0.5)),
    )

    if args.lookup:
        client.lookup_corp_code(args.lookup)
        return 0

    target_date = parse_target_date(args.date)
    yyyymmdd = to_yyyymmdd(target_date)
    yymmdd = to_yymmdd(target_date)

    classification = cfg.get("classification", {})
    pblntf_types = dart_cfg.get("pblntf_types", ["A", "B", "D", "E", "I"])
    watchlist_cfg = cfg.get("watchlist", [])

    print(f"[시작] {title_date_text(target_date)} DART 공시 수집")

    # 1) 전체 시장 공시 수집
    print(f"[수집] 전체 시장 공시 (유형: {', '.join(pblntf_types)})")
    raw_market = client.fetch_all_types(
        bgn_de=yyyymmdd,
        end_de=yyyymmdd,
        pblntf_types=pblntf_types,
    )
    market_rows = parse_disclosure_rows(raw_market, classification)
    print(f"  → {len(market_rows)}건 수집")

    # 2) 관심종목 공시 수집
    watchlist_rows: list[DisclosureRow] = []
    if watchlist_cfg:
        names = ", ".join(w.get("name", "") for w in watchlist_cfg)
        print(f"[수집] 관심종목 ({names})")
        raw_watchlist = client.fetch_watchlist(
            bgn_de=yyyymmdd,
            end_de=yyyymmdd,
            watchlist=watchlist_cfg,
        )
        watchlist_rows = parse_disclosure_rows(raw_watchlist, classification)
        print(f"  → {len(watchlist_rows)}건 수집")

    # 3) 요약 집계
    summaries = build_summaries(market_rows, SHEET_TYPE_MAP)

    # 4) 시트별 그룹핑
    by_sheet = group_by_sheet(market_rows, SHEET_TYPE_MAP)

    # 5) 시세 데이터 로드 (KRX 시세 파일에서: 시총 + 등락률)
    output_dir = Path(cfg.get("output_dir", "시장 정리")).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    market_cap_by_code, pct_by_code = load_price_data(output_dir, yymmdd)
    if market_cap_by_code:
        print(f"[시세] {len(market_cap_by_code)}개 종목 시세 데이터 로드 (시총+등락률)")
    else:
        print("[시세] 시세 데이터 없음 (시총/등락률 컬럼 비어있음)")

    # 6) 엑셀 출력
    out_path = output_dir / f"{yymmdd}_공시.xlsx"

    write_dart_workbook(
        out_path=out_path,
        target_date=target_date,
        summaries=summaries,
        by_sheet=by_sheet,
        watchlist=watchlist_rows,
        market_cap_by_code=market_cap_by_code,
        pct_by_code=pct_by_code,
    )

    print(f"[완료] {out_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"[오류] {exc}", file=sys.stderr)
        raise
