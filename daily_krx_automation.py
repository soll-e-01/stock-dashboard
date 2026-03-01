from __future__ import annotations

import argparse
import csv
import io
import json
import sys
import time
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import requests
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]
INVESTOR_SHEETS = ["사모펀드", "투자신탁", "연기금", "외국인", "개인"]
HIGH_SHEETS = ["역대 신고가", "연중 신고가", "52주 신고가"]
DATE_TOKEN = "__DATE__"

# ---------------------------------------------------------------------------
# KSIC 업종코드 → 한줄소개 매핑
# ---------------------------------------------------------------------------

KSIC_SHORT: dict[str, str] = {
    "01": "작물재배", "02": "축산", "03": "임업", "05": "광업",
    "10": "식료품", "11": "음료", "13": "섬유", "14": "의류/패션",
    "15": "가죽/신발", "17": "제지", "19": "석유정제",
    "20": "화학", "21": "바이오/제약",
    "22": "고무/플라스틱", "23": "비금속광물",
    "24": "철강/금속", "25": "금속가공",
    "26": "전자/IT",
    "261": "반도체", "262": "전자부품", "263": "컴퓨터",
    "264": "통신장비", "265": "영상/음향",
    "27": "의료기기/정밀",
    "271": "의료기기", "272": "측정/정밀기기",
    "28": "전기장비", "282": "2차전지",
    "29": "기계장비",
    "30": "자동차", "303": "자동차부품",
    "31": "운송장비", "311": "조선", "313": "항공우주",
    "33": "기타제조",
    "35": "전기/가스", "41": "건설", "42": "전문건설",
    "46": "도매/유통", "47": "소매/유통",
    "49": "육상운송", "50": "해운", "51": "항공",
    "52": "물류/창고",
    "58": "출판/미디어", "582": "소프트웨어",
    "59": "영상/콘텐츠", "60": "방송",
    "61": "통신", "62": "IT서비스",
    "63": "인터넷/플랫폼",
    "64": "금융/은행", "65": "보험", "66": "금융서비스",
    "68": "부동산", "70": "연구개발",
    "71": "전문서비스", "72": "엔지니어링",
    "74": "시설관리", "75": "사업지원",
    "86": "의료/헬스케어",
    "90": "엔터/여가", "91": "스포츠/레저",
}


def get_industry_short(induty_code: str) -> str:
    """KSIC 코드에서 짧은 업종명 반환 (3자리 우선, 2자리 폴백)."""
    code = induty_code.strip()
    for length in (3, 2):
        if len(code) >= length and code[:length] in KSIC_SHORT:
            return KSIC_SHORT[code[:length]]
    return ""


def fetch_company_industries(
    stock_codes: set[str],
    api_key: str,
    base_url: str = "https://opendart.fss.or.kr/api",
    timeout: int = 30,
    cache_path: Path | None = None,
) -> dict[str, str]:
    """종목코드별 업종 한줄소개 반환 (DART API + 캐시)."""
    cache: dict[str, str] = {}
    if cache_path and cache_path.exists():
        try:
            with cache_path.open("r", encoding="utf-8") as f:
                cache = json.load(f)
        except Exception:
            cache = {}

    needed = {c for c in stock_codes if c not in cache}
    if not needed:
        return {c: cache.get(c, "") for c in stock_codes}

    # corpCode.xml 다운로드 → stock_code → corp_code 매핑
    print(f"[업종] DART corpCode.xml 다운로드 중 ({len(needed)}개 조회)...")
    try:
        session = requests.Session()
        resp = session.get(
            f"{base_url}/corpCode.xml",
            params={"crtfc_key": api_key},
            timeout=60,
        )
        resp.raise_for_status()
        with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
            xml_name = zf.namelist()[0]
            root = ET.fromstring(zf.read(xml_name))
        corp_map: dict[str, str] = {}
        for item in root.iter("list"):
            sc = (item.findtext("stock_code") or "").strip()
            cc = (item.findtext("corp_code") or "").strip()
            if sc and cc:
                corp_map[sc] = cc
    except Exception as exc:
        print(f"[업종] corpCode.xml 다운로드 실패: {exc}")
        return {c: cache.get(c, "") for c in stock_codes}

    fetched = 0
    for code in needed:
        corp_code = corp_map.get(code, "")
        if not corp_code:
            cache[code] = ""
            continue
        try:
            resp = session.get(
                f"{base_url}/company.json",
                params={"crtfc_key": api_key, "corp_code": corp_code},
                timeout=timeout,
            )
            resp.raise_for_status()
            body = resp.json()
            if body.get("status") == "000":
                induty_code = body.get("induty_code", "")
                cache[code] = get_industry_short(induty_code)
                fetched += 1
            else:
                cache[code] = ""
            time.sleep(0.3)
        except Exception:
            cache[code] = ""

    print(f"[업종] {fetched}개 종목 업종 정보 수집 완료")

    if cache_path:
        try:
            with cache_path.open("w", encoding="utf-8") as f:
                json.dump(cache, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    return {c: cache.get(c, "") for c in stock_codes}


def name_with_industry(name: str, code: str, industries: dict[str, str]) -> str:
    """종목명에 업종 정보 병기 (예: '삼성전자 (반도체)')."""
    ind = industries.get(code, "")
    if ind:
        return f"{name} ({ind})"
    return name

DEFAULT_PRICE_FIELD_MAP: dict[str, list[str]] = {
    "code": ["ISU_SRT_CD", "종목코드"],
    "name": ["ISU_ABBRV", "종목명"],
    "market": ["MKT_NM", "시장구분", "시장"],
    "sector": ["SECT_TP_NM", "소속부"],
    "close": ["TDD_CLSPRC", "종가"],
    "change": ["CMPPREVDD_PRC", "대비"],
    "pct": ["FLUC_RT", "등락률", "등락률(%)"],
    "open": ["TDD_OPNPRC", "시가"],
    "high": ["TDD_HGPRC", "고가"],
    "low": ["TDD_LWPRC", "저가"],
    "volume": ["ACC_TRDVOL", "거래량"],
    "trade_value": ["ACC_TRDVAL", "거래대금"],
    "market_cap": ["MKTCAP", "시가총액"],
    "listed_shares": ["LIST_SHRS", "상장주식수"],
}

DEFAULT_SUPPLY_FIELD_MAP: dict[str, list[str]] = {
    "code": ["ISU_SRT_CD", "종목코드"],
    "name": ["ISU_NM", "ISU_ABBRV", "종목명"],
    "market": ["MKT_NM", "시장", "시장구분"],
    "net_buy": ["NETBID_TRDVAL", "순매수거래대금", "순매수 거래대금"],
    "market_cap": ["MKTCAP", "시가총액"],
}

DEFAULT_HIGH_FIELD_MAP: dict[str, list[str]] = {
    "code": ["ISU_CD", "ISU_SRT_CD", "종목코드"],
    "name": ["ISU_ABBRV", "종목명"],
    "market": ["MKT_NM", "시장", "시장구분"],
    "market_cap": ["MKTCAP", "시가총액(억)", "시가총액"],
    "pct": ["FLUC_RT", "당일상승률(%)", "등락률", "등락률(%)"],
    "high_price": ["TDD_CLSPRC", "당일고가", "고가"],
}


@dataclass
class PriceRow:
    code: str
    name: str
    market: str
    sector: str
    close: int
    change: int
    pct: float
    open_price: int
    high: int
    low: int
    volume: int
    trade_value: int
    market_cap: int
    listed_shares: int


@dataclass
class SupplyRow:
    code: str
    name: str
    market: str
    net_buy: int
    market_cap: int

    @property
    def ratio(self) -> float:
        if self.market_cap <= 0:
            return 0.0
        return self.net_buy / self.market_cap


@dataclass
class HighRow:
    code: str
    name: str
    market: str
    market_cap_eok: int
    pct: float
    high_price: int


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="KRX 일일 데이터 수집 후 3종 엑셀(수급/시세/신고가) 생성"
    )
    p.add_argument(
        "--config",
        default="config.krx.json",
        help="설정 파일 경로 (기본: config.krx.json)",
    )
    p.add_argument(
        "--date",
        default="",
        help="대상 일자 YYYYMMDD. 미입력 시 오늘 날짜 사용",
    )
    p.add_argument(
        "--dart-config",
        default="",
        help="DART 설정 파일 (업종 정보 조회용, 예: config.dart.json)",
    )
    return p.parse_args()


def parse_target_date(raw: str) -> date:
    if raw:
        return datetime.strptime(raw, "%Y%m%d").date()
    return date.today()


def find_last_business_day(d: date) -> date:
    """주말이면 직전 영업일(금요일)로 조정."""
    while d.weekday() >= 5:  # 5=토, 6=일
        d -= timedelta(days=1)
    return d


def to_yyyymmdd(d: date) -> str:
    return d.strftime("%Y%m%d")


def to_yymmdd(d: date) -> str:
    return d.strftime("%y%m%d")


def title_date_text(d: date) -> str:
    return f"{d:%Y/%m/%d}({WEEKDAY_KR[d.weekday()]})"


def as_text_code(v: Any) -> str:
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if s.isdigit():
        return s.zfill(6)
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits and len(digits) <= 6:
        return digits.zfill(6)
    return s


def as_float(v: Any) -> float:
    if v is None:
        return 0.0
    s = str(v).strip()
    if not s:
        return 0.0
    s = s.replace(",", "").replace("%", "").replace("+", "")
    if s in {"-", "--"}:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def as_int(v: Any) -> int:
    return int(round(as_float(v)))


def as_percent_text(ratio: float) -> str:
    return f"{ratio * 100:.1f}%"


def as_eok_from_won(v: int) -> int:
    return int(round(v / 100_000_000))


def ensure_file(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"{label} 파일이 없습니다: {path}")


def create_price_template(path: Path) -> None:
    """시세 템플릿 엑셀 자동 생성 (최초 실행용)."""
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "DATA"
    headers = [
        "종목코드", "종목명", "시장구분", "소속부", "종가", "대비",
        "등락률", "시가", "고가", "저가", "거래량", "거래대금",
        "시가총액", "상장주식수",
    ]
    for c, h in enumerate(headers, 1):
        ws_data.cell(1, c, h)
    for name in ["등락률 상위", "거래대금 상위", "시총대비 거래대금 상위"]:
        ws = wb.create_sheet(name)
        ws.cell(1, 1, "")
        ws.cell(2, 1, "")
        ws.cell(3, 1, "No.")
        ws.cell(3, 2, "종목명")
        ws.cell(3, 3, "시가총액(억)")
        ws.cell(3, 4, "거래대금(억)")
        ws.cell(3, 5, "등락률(%)")
    wb.save(path)
    print(f"[템플릿] 시세 템플릿 자동 생성: {path}")


def create_supply_template(path: Path) -> None:
    """수급 템플릿 엑셀 자동 생성 (최초 실행용)."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "수급"
    ws_summary.cell(1, 1, "")
    ws_summary.cell(2, 1, "")
    for name in INVESTOR_SHEETS:
        ws = wb.create_sheet(name)
        ws.cell(1, 1, f"{name} 원본 데이터")
        ws.cell(2, 1, "종목코드")
        ws.cell(2, 2, "종목명")
        ws.cell(2, 3, "시장")
        ws.cell(2, 4, "순매수거래대금")
        ws.cell(2, 5, "시가총액")
        ws.cell(2, 6, "시총대비순매수")
    wb.save(path)
    print(f"[템플릿] 수급 템플릿 자동 생성: {path}")


def create_high_template(path: Path) -> None:
    """신고가 템플릿 엑셀 자동 생성 (최초 실행용)."""
    wb = Workbook()
    wb.remove(wb.active)
    for name in HIGH_SHEETS:
        ws = wb.create_sheet(name)
        ws.cell(1, 1, "No.")
        ws.cell(1, 2, "종목코드")
        ws.cell(1, 3, "종목명")
        ws.cell(1, 4, "시장")
        ws.cell(1, 5, "시가총액(억)")
        ws.cell(1, 6, "당일상승률(%)")
        ws.cell(1, 7, "당일고가")
    wb.save(path)
    print(f"[템플릿] 신고가 템플릿 자동 생성: {path}")


def ensure_sheet(wb: Workbook, sheet_name: str) -> Any:
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"시트를 찾을 수 없습니다: {sheet_name}")
    return wb[sheet_name]


def clear_rect(ws: Any, start_row: int, end_row: int, end_col: int) -> None:
    if end_row < start_row:
        return
    for r in range(start_row, end_row + 1):
        for c in range(1, end_col + 1):
            ws.cell(r, c, None)


def replace_tokens(data: Any, yyyymmdd: str) -> Any:
    if isinstance(data, dict):
        return {k: replace_tokens(v, yyyymmdd) for k, v in data.items()}
    if isinstance(data, list):
        return [replace_tokens(v, yyyymmdd) for v in data]
    if isinstance(data, str):
        return data.replace(DATE_TOKEN, yyyymmdd)
    return data


def resolve_field_candidates(
    src_row: dict[str, str],
    field_name: str,
    candidates: list[str],
) -> str:
    for col in candidates:
        if col in src_row:
            return src_row[col]
    cols = ", ".join(src_row.keys())
    cand = ", ".join(candidates)
    raise KeyError(f"필드 매핑 실패: {field_name} -> [{cand}] / 실제컬럼=[{cols}]")


def normalize_field_map(raw_map: dict[str, Any], default_map: dict[str, list[str]]) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    for key, default_candidates in default_map.items():
        value = raw_map.get(key, default_candidates)
        if isinstance(value, str):
            out[key] = [value]
        elif isinstance(value, list) and value:
            out[key] = [str(x) for x in value]
        else:
            out[key] = default_candidates
    return out


class KrxOtpClient:
    def __init__(
        self,
        otp_url: str,
        download_url: str,
        json_url: str,
        headers: dict[str, str],
        timeout_sec: int,
        encoding: str,
        login_page_url: str = "",
        login_url: str = "",
        credentials: dict[str, str] | None = None,
    ):
        self.otp_url = otp_url
        self.download_url = download_url
        self.json_url = json_url
        self.timeout_sec = timeout_sec
        self.encoding = encoding
        self.login_page_url = login_page_url
        self.login_url = login_url
        self.credentials = credentials or {}
        self.session = requests.Session()
        self.session.headers.update(headers)

    def login(self) -> None:
        if not self.login_url or not self.credentials:
            return
        if self.login_page_url:
            self.session.get(self.login_page_url, timeout=self.timeout_sec)
        resp = self.session.post(
            self.login_url,
            data=self.credentials,
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "Referer": self.login_page_url,
            },
            timeout=self.timeout_sec,
        )
        resp.raise_for_status()
        body = resp.json()
        err = body.get("_error_code", "")
        if err:
            msg = body.get("_error_message", "알 수 없는 오류")
            if err == "CD001":
                print("[로그인] KRX 로그인 성공")
            elif err == "CD011":
                print("[로그인] 중복 로그인 감지 - 기존 세션으로 계속 진행")
            else:
                raise RuntimeError(f"KRX 로그인 실패 [{err}]: {msg}")
        else:
            print("[로그인] KRX 로그인 성공")

    def fetch_rows(self, otp_params: dict[str, Any]) -> list[dict[str, str]]:
        otp = self._request_otp(otp_params)
        raw = self._download_csv_bytes(otp)
        text = self._decode_csv(raw)
        return self._parse_csv(text)

    def fetch_rows_json(self, request_params: dict[str, Any]) -> list[dict[str, str]]:
        resp = self.session.post(
            self.json_url,
            data={k: str(v) for k, v in request_params.items()},
            timeout=self.timeout_sec,
        )
        resp.raise_for_status()
        return self._parse_json_rows(resp.content)

    def _request_otp(self, otp_params: dict[str, Any]) -> str:
        resp = self.session.post(
            self.otp_url,
            data={k: str(v) for k, v in otp_params.items()},
            timeout=self.timeout_sec,
        )
        resp.raise_for_status()
        otp = resp.text.strip()
        if not otp or otp == "LOGOUT":
            raise RuntimeError("OTP 발급 실패: 빈 응답 또는 LOGOUT")
        return otp

    def _download_csv_bytes(self, otp: str) -> bytes:
        resp = self.session.post(
            self.download_url,
            data={"code": otp},
            timeout=self.timeout_sec,
        )
        resp.raise_for_status()
        return resp.content

    def _decode_csv(self, raw: bytes) -> str:
        for enc in [self.encoding, "cp949", "euc-kr", "utf-8-sig", "utf-8"]:
            try:
                return raw.decode(enc)
            except UnicodeDecodeError:
                continue
        raise UnicodeDecodeError("csv", b"", 0, 0, "CSV 디코딩 실패")

    @staticmethod
    def _parse_csv(text: str) -> list[dict[str, str]]:
        reader = csv.DictReader(io.StringIO(text))
        out: list[dict[str, str]] = []
        for row in reader:
            clean: dict[str, str] = {}
            for k, v in row.items():
                key = (k or "").strip()
                if not key:
                    continue
                clean[key] = (v or "").strip()
            if clean:
                out.append(clean)
        return out

    def _parse_json_rows(self, raw: bytes) -> list[dict[str, str]]:
        payload = self._decode_json_text(raw)
        try:
            obj = json.loads(payload)
        except json.JSONDecodeError as exc:
            raise RuntimeError(f"JSON 파싱 실패: {exc}") from exc

        rows = self._extract_rows_from_json(obj)
        out: list[dict[str, str]] = []
        for row in rows:
            clean: dict[str, str] = {}
            for k, v in row.items():
                key = str(k).strip()
                if not key:
                    continue
                clean[key] = "" if v is None else str(v).strip()
            if clean:
                out.append(clean)
        if not out:
            raise RuntimeError("JSON 응답에서 데이터 행을 찾지 못했습니다.")
        return out

    def _decode_json_text(self, raw: bytes) -> str:
        for enc in ["utf-8-sig", "utf-8", self.encoding, "cp949", "euc-kr"]:
            try:
                return raw.decode(enc)
            except UnicodeDecodeError:
                continue
        raise UnicodeDecodeError("json", b"", 0, 0, "JSON 디코딩 실패")

    @classmethod
    def _extract_rows_from_json(cls, obj: Any) -> list[dict[str, Any]]:
        if isinstance(obj, list):
            if obj and all(isinstance(x, dict) for x in obj):
                return obj
            for item in obj:
                rows = cls._extract_rows_from_json(item)
                if rows:
                    return rows
            return []

        if isinstance(obj, dict):
            preferred_keys = [
                "OutBlock_1",
                "outBlock_1",
                "OutBlock1",
                "outBlock1",
                "output",
                "result",
            ]
            for key in preferred_keys:
                if key in obj:
                    rows = cls._extract_rows_from_json(obj[key])
                    if rows:
                        return rows
            for value in obj.values():
                rows = cls._extract_rows_from_json(value)
                if rows:
                    return rows
        return []


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def parse_price_rows(src_rows: list[dict[str, str]], field_map: dict[str, list[str]]) -> list[PriceRow]:
    out: list[PriceRow] = []
    for row in src_rows:
        out.append(
            PriceRow(
                code=as_text_code(resolve_field_candidates(row, "code", field_map["code"])),
                name=resolve_field_candidates(row, "name", field_map["name"]),
                market=resolve_field_candidates(row, "market", field_map["market"]),
                sector=resolve_field_candidates(row, "sector", field_map["sector"]),
                close=as_int(resolve_field_candidates(row, "close", field_map["close"])),
                change=as_int(resolve_field_candidates(row, "change", field_map["change"])),
                pct=as_float(resolve_field_candidates(row, "pct", field_map["pct"])),
                open_price=as_int(resolve_field_candidates(row, "open", field_map["open"])),
                high=as_int(resolve_field_candidates(row, "high", field_map["high"])),
                low=as_int(resolve_field_candidates(row, "low", field_map["low"])),
                volume=as_int(resolve_field_candidates(row, "volume", field_map["volume"])),
                trade_value=as_int(resolve_field_candidates(row, "trade_value", field_map["trade_value"])),
                market_cap=as_int(resolve_field_candidates(row, "market_cap", field_map["market_cap"])),
                listed_shares=as_int(resolve_field_candidates(row, "listed_shares", field_map["listed_shares"])),
            )
        )
    return out


def parse_supply_rows(src_rows: list[dict[str, str]], field_map: dict[str, list[str]]) -> list[SupplyRow]:
    out: list[SupplyRow] = []
    for row in src_rows:
        out.append(
            SupplyRow(
                code=as_text_code(resolve_field_candidates(row, "code", field_map["code"])),
                name=resolve_field_candidates(row, "name", field_map["name"]),
                market=resolve_field_candidates(row, "market", field_map["market"]),
                net_buy=as_int(resolve_field_candidates(row, "net_buy", field_map["net_buy"])),
                market_cap=as_int(resolve_field_candidates(row, "market_cap", field_map["market_cap"])),
            )
        )
    return out


def parse_high_rows(
    src_rows: list[dict[str, str]],
    field_map: dict[str, list[str]],
    market_cap_unit: str,
) -> list[HighRow]:
    out: list[HighRow] = []
    for row in src_rows:
        market_cap_raw = as_int(resolve_field_candidates(row, "market_cap", field_map["market_cap"]))
        if market_cap_unit.lower() == "won":
            market_cap_eok = as_eok_from_won(market_cap_raw)
        else:
            market_cap_eok = market_cap_raw
        out.append(
            HighRow(
                code=as_text_code(resolve_field_candidates(row, "code", field_map["code"])),
                name=resolve_field_candidates(row, "name", field_map["name"]),
                market=resolve_field_candidates(row, "market", field_map["market"]),
                market_cap_eok=market_cap_eok,
                pct=as_float(resolve_field_candidates(row, "pct", field_map["pct"])),
                high_price=as_int(resolve_field_candidates(row, "high_price", field_map["high_price"])),
            )
        )
    return out


def top_n(items: list[Any], key_fn: Any, n: int = 30) -> list[Any]:
    return sorted(items, key=key_fn, reverse=True)[:n]


def write_price_workbook(
    template_path: Path,
    out_path: Path,
    target_date: date,
    prices: list[PriceRow],
    industries: dict[str, str] | None = None,
) -> None:
    wb = load_workbook(template_path)
    ws_data = ensure_sheet(wb, "DATA")
    ws_top_pct = ensure_sheet(wb, "등락률 상위")
    ws_top_value = ensure_sheet(wb, "거래대금 상위")
    ws_top_ratio = ensure_sheet(wb, "시총대비 거래대금 상위")

    old_rows = ws_data.max_row
    new_last_row = max(2, len(prices) + 1)
    clear_rect(ws_data, 2, max(old_rows, new_last_row), 14)
    for idx, p in enumerate(prices, start=2):
        ws_data.cell(idx, 1, p.code)
        ws_data.cell(idx, 2, p.name)
        ws_data.cell(idx, 3, p.market)
        ws_data.cell(idx, 4, p.sector)
        ws_data.cell(idx, 5, p.close)
        ws_data.cell(idx, 6, p.change)
        ws_data.cell(idx, 7, p.pct)
        ws_data.cell(idx, 8, p.open_price)
        ws_data.cell(idx, 9, p.high)
        ws_data.cell(idx, 10, p.low)
        ws_data.cell(idx, 11, p.volume)
        ws_data.cell(idx, 12, p.trade_value)
        ws_data.cell(idx, 13, p.market_cap)
        ws_data.cell(idx, 14, p.listed_shares)

    for table in ws_data.tables.values():
        table.ref = f"A1:N{new_last_row}"

    dt_text = title_date_text(target_date)
    ws_top_pct["A1"] = f"{dt_text} 등락률 상위 종목"
    ws_top_pct["A2"] = "등락률 상위 Top 30"
    ws_top_value["A1"] = f"{dt_text} 거래대금 상위"
    ws_top_value["A2"] = "거래대금 상위 Top 30"
    ws_top_ratio["A1"] = f"{dt_text} 시총대비 거래대금 상위"
    ws_top_ratio["A2"] = "시총대비 거래대금 상위 Top 30"

    top_pct = top_n(prices, key_fn=lambda x: x.pct, n=30)
    top_value = top_n(prices, key_fn=lambda x: x.trade_value, n=30)
    top_ratio = top_n(
        [x for x in prices if x.market_cap > 0],
        key_fn=lambda x: (x.trade_value / x.market_cap),
        n=30,
    )

    fill_price_top_sheet(ws_top_pct, top_pct, industries)
    fill_price_top_sheet(ws_top_value, top_value, industries)
    fill_price_top_sheet(ws_top_ratio, top_ratio, industries)

    wb.save(out_path)


def fill_price_top_sheet(ws: Any, rows: list[PriceRow], industries: dict[str, str] | None = None) -> None:
    ind = industries or {}
    clear_rect(ws, 4, max(ws.max_row, 33), 5)
    for i, r in enumerate(rows, start=1):
        rr = i + 3
        ws.cell(rr, 1, i)
        ws.cell(rr, 2, name_with_industry(r.name, r.code, ind))
        ws.cell(rr, 3, as_eok_from_won(r.market_cap))
        ws.cell(rr, 4, as_eok_from_won(r.trade_value))
        ws.cell(rr, 5, round(r.pct, 2))


def write_supply_workbook(
    template_path: Path,
    out_path: Path,
    target_date: date,
    supply_rows_by_sheet: dict[str, list[SupplyRow]],
    industries: dict[str, str] | None = None,
) -> None:
    wb = load_workbook(template_path)
    ws_summary = ensure_sheet(wb, "수급")

    dt_text = title_date_text(target_date)
    ws_summary["A1"] = f"{dt_text} 수급 상위 종목"
    ws_summary["A2"] = "순매수 금액 상위 10 / 시총대비 순매수 상위 10 (코스피+코스닥)"

    for name in INVESTOR_SHEETS:
        ws = ensure_sheet(wb, name)
        rows = sorted(supply_rows_by_sheet.get(name, []), key=lambda x: x.net_buy, reverse=True)
        clear_rect(ws, 1, max(ws.max_row, len(rows) + 2), 6)
        ws.cell(1, 1, f"{name} 원본 데이터")
        ws.cell(2, 1, "종목코드")
        ws.cell(2, 2, "종목명")
        ws.cell(2, 3, "시장")
        ws.cell(2, 4, "순매수거래대금")
        ws.cell(2, 5, "시가총액")
        ws.cell(2, 6, "시총대비순매수")

        for idx, r in enumerate(rows, start=3):
            ws.cell(idx, 1, r.code)
            ws.cell(idx, 2, r.name)
            ws.cell(idx, 3, r.market)
            ws.cell(idx, 4, r.net_buy)
            ws.cell(idx, 5, r.market_cap)
            ws.cell(idx, 6, r.ratio)

    fill_supply_summary(ws_summary, supply_rows_by_sheet, industries)
    wb.save(out_path)


def fill_supply_summary(ws: Any, supply_rows_by_sheet: dict[str, list[SupplyRow]], industries: dict[str, str] | None = None) -> None:
    ind = industries or {}
    clear_rect(ws, 4, max(ws.max_row, 67), 10)
    start_rows = [4, 17, 30, 43, 56]
    for sheet_name, base in zip(INVESTOR_SHEETS, start_rows):
        rows = supply_rows_by_sheet.get(sheet_name, [])
        top_value = top_n(rows, key_fn=lambda x: x.net_buy, n=10)
        top_ratio = top_n(rows, key_fn=lambda x: x.ratio, n=10)

        ws.cell(base, 1, f"# {sheet_name}")
        ws.cell(base, 6, f"# {sheet_name}")

        ws.cell(base + 1, 1, "No.")
        ws.cell(base + 1, 2, "종목명")
        ws.cell(base + 1, 3, "순매수 금액(억)")
        ws.cell(base + 1, 4, "시가총액(억)")
        ws.cell(base + 1, 6, "No.")
        ws.cell(base + 1, 7, "종목명")
        ws.cell(base + 1, 8, "시총대비 순매수(%)")
        ws.cell(base + 1, 9, "순매수 금액(억)")
        ws.cell(base + 1, 10, "시가총액(억)")

        for i in range(10):
            rr = base + 2 + i
            if i < len(top_value):
                lv = top_value[i]
                ws.cell(rr, 1, i + 1)
                ws.cell(rr, 2, name_with_industry(lv.name, lv.code, ind))
                ws.cell(rr, 3, as_eok_from_won(lv.net_buy))
                ws.cell(rr, 4, as_eok_from_won(lv.market_cap))
            if i < len(top_ratio):
                rv = top_ratio[i]
                ws.cell(rr, 6, i + 1)
                ws.cell(rr, 7, name_with_industry(rv.name, rv.code, ind))
                ws.cell(rr, 8, as_percent_text(rv.ratio))
                ws.cell(rr, 9, as_eok_from_won(rv.net_buy))
                ws.cell(rr, 10, as_eok_from_won(rv.market_cap))


def write_high_workbook(
    template_path: Path,
    out_path: Path,
    high_rows_by_sheet: dict[str, list[HighRow]],
    industries: dict[str, str] | None = None,
) -> None:
    ind = industries or {}
    wb = load_workbook(template_path)

    # 기존 템플릿 시트 중 데이터가 없는 시트 삭제
    for name in list(wb.sheetnames):
        if name not in high_rows_by_sheet:
            del wb[name]

    for name, rows in high_rows_by_sheet.items():
        if name in wb.sheetnames:
            ws = wb[name]
        else:
            ws = wb.create_sheet(name)
        clear_rect(ws, 1, max(ws.max_row, len(rows) + 1), 7)
        ws.cell(1, 1, "No.")
        ws.cell(1, 2, "종목코드")
        ws.cell(1, 3, "종목명")
        ws.cell(1, 4, "시장")
        ws.cell(1, 5, "시가총액(억)")
        ws.cell(1, 6, "당일상승률(%)")
        ws.cell(1, 7, "당일고가")
        for idx, r in enumerate(rows, start=2):
            ws.cell(idx, 1, idx - 1)
            ws.cell(idx, 2, r.code)
            ws.cell(idx, 3, name_with_industry(r.name, r.code, ind))
            ws.cell(idx, 4, r.market)
            ws.cell(idx, 5, r.market_cap_eok)
            ws.cell(idx, 6, round(r.pct, 2))
            ws.cell(idx, 7, r.high_price)
    wb.save(out_path)


def read_query_rows(
    client: KrxOtpClient,
    query_cfg: dict[str, Any],
    yyyymmdd: str,
) -> list[dict[str, str]]:
    if "request_params" in query_cfg:
        request_params = replace_tokens(query_cfg["request_params"], yyyymmdd)
        return client.fetch_rows_json(request_params=request_params)
    if "otp_params" in query_cfg:
        otp_params = replace_tokens(query_cfg["otp_params"], yyyymmdd)
        return client.fetch_rows(otp_params=otp_params)
    raise KeyError("query 설정에 request_params 또는 otp_params가 필요합니다.")


def fetch_with_retry(
    client: KrxOtpClient,
    query_cfg: dict[str, Any],
    yyyymmdd: str,
    label: str = "",
    max_retries: int = 3,
    base_delay: float = 2.0,
) -> list[dict[str, str]]:
    """재시도 로직이 포함된 데이터 조회."""
    last_err: Exception | None = None
    for attempt in range(1, max_retries + 1):
        try:
            return read_query_rows(client, query_cfg, yyyymmdd)
        except Exception as e:
            last_err = e
            if attempt < max_retries:
                wait = base_delay * (1.5 ** (attempt - 1))
                print(f"  [{label}] 시도 {attempt}/{max_retries} 실패: {e}")
                print(f"  [{label}] {wait:.1f}초 후 재시도...")
                time.sleep(wait)
    raise RuntimeError(f"[{label}] {max_retries}회 시도 모두 실패: {last_err}")


def build_client(cfg: dict[str, Any]) -> KrxOtpClient:
    krx = cfg["krx"]
    otp_url = krx.get("otp_url", "https://data.krx.co.kr/comm/fileDn/GenerateOTP/generate.cmd")
    download_url = krx.get("download_url", "https://data.krx.co.kr/comm/fileDn/download_csv/download.cmd")
    json_url = krx.get("json_url", "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd")
    headers = krx.get("headers", {})
    timeout_sec = int(krx.get("timeout_sec", 30))
    encoding = str(krx.get("encoding", "cp949"))
    login_cfg = krx.get("login", {})
    login_page_url = krx.get("login_page_url", "")
    login_url = krx.get("login_url", "")
    return KrxOtpClient(
        otp_url=otp_url,
        download_url=download_url,
        json_url=json_url,
        headers=headers,
        timeout_sec=timeout_sec,
        encoding=encoding,
        login_page_url=login_page_url,
        login_url=login_url,
        credentials=login_cfg if login_cfg else None,
    )


def main() -> int:
    args = parse_args()
    config_path = Path(args.config).resolve()
    ensure_file(config_path, "설정")
    cfg = load_json(config_path)

    target_date = parse_target_date(args.date)
    original_date = target_date
    target_date = find_last_business_day(target_date)
    if target_date != original_date:
        print(f"[날짜] {to_yyyymmdd(original_date)}은(는) 주말이므로 직전 영업일 {to_yyyymmdd(target_date)} 사용")
    yyyymmdd = to_yyyymmdd(target_date)
    yymmdd = to_yymmdd(target_date)

    template_cfg = cfg.get("template_files", {})
    output_dir = Path(cfg.get("output_dir", ".")).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    template_price = Path(template_cfg.get("price", "260213_시세_DATA.xlsx")).resolve()
    template_supply = Path(template_cfg.get("supply", "260213_수급_DATA.xlsx")).resolve()
    template_high = Path(template_cfg.get("high", "260213_신고가.xlsx")).resolve()
    if not template_price.exists():
        create_price_template(template_price)
    if not template_supply.exists():
        create_supply_template(template_supply)
    if not template_high.exists():
        create_high_template(template_high)

    # ---------------------------------------------------------------
    # KRX 직접 API로 데이터 조회
    # ---------------------------------------------------------------
    client = build_client(cfg)

    # 세션 초기화 (메인 페이지 방문으로 쿠키 획득)
    init_url = cfg["krx"].get(
        "login_page_url",
        "https://data.krx.co.kr/contents/MDC/MDI/mdiLoader/index.cmd",
    )
    try:
        client.session.get(init_url, timeout=client.timeout_sec)
        print("[KRX] 세션 쿠키 획득")
    except Exception as e:
        print(f"[KRX] 세션 초기화 경고(계속 진행): {e}")

    client.login()

    datasets = cfg.get("datasets", {})

    # --- 1) 시세(전종목) 조회 ---
    price_cfg = datasets.get("price_all", {})
    price_fm = normalize_field_map(
        price_cfg.get("field_map", {}), DEFAULT_PRICE_FIELD_MAP,
    )
    print(f"[KRX] {yyyymmdd} 시세 데이터 조회 중...")
    raw_prices = fetch_with_retry(client, price_cfg, yyyymmdd, "시세")
    price_rows = parse_price_rows(raw_prices, price_fm)
    if not price_rows:
        raise RuntimeError(f"{yyyymmdd} 시세 데이터가 없습니다 (휴장일?)")
    print(f"[KRX] 시세: {len(price_rows)}종목")

    # 시세 데이터에서 룩업맵 생성 (수급·신고가 보강용)
    mcap_map: dict[str, int] = {r.code: r.market_cap for r in price_rows}
    mkt_map: dict[str, str] = {r.code: r.market for r in price_rows}
    time.sleep(1)

    # --- 2) 투자자별 수급 조회 ---
    supply_cfgs = datasets.get("supply", {})
    supply_fm = normalize_field_map({}, DEFAULT_SUPPLY_FIELD_MAP)
    supply_rows_by_sheet: dict[str, list[SupplyRow]] = {}
    for sheet_name in INVESTOR_SHEETS:
        inv_cfg = supply_cfgs.get(sheet_name)
        if not inv_cfg:
            print(f"[KRX] 수급({sheet_name}) 설정 없음 - 건너뜀")
            supply_rows_by_sheet[sheet_name] = []
            continue
        print(f"[KRX] 수급({sheet_name}) 조회 중...")
        try:
            raw = fetch_with_retry(client, inv_cfg, yyyymmdd, f"수급-{sheet_name}")
            # 시가총액·시장 필드가 없으면 시세 데이터로 보강
            for row in raw:
                code = as_text_code(
                    row.get("ISU_SRT_CD", row.get("종목코드", ""))
                )
                if not any(f in row for f in supply_fm["market_cap"]):
                    row["시가총액"] = str(mcap_map.get(code, 0))
                if not any(f in row for f in supply_fm["market"]):
                    row["시장구분"] = mkt_map.get(code, "")
            rows = parse_supply_rows(raw, supply_fm)
            supply_rows_by_sheet[sheet_name] = rows
            print(f"  → {len(rows)}종목")
        except Exception as e:
            print(f"  → 실패: {e}")
            supply_rows_by_sheet[sheet_name] = []
        time.sleep(1)

    # --- 3) 신고가 조회 ---
    high_cfgs = datasets.get("highs", {})
    high_fm = normalize_field_map({}, DEFAULT_HIGH_FIELD_MAP)
    market_cap_unit = str(high_cfgs.get("market_cap_unit", "won"))
    high_rows_by_sheet: dict[str, list[HighRow]] = {}
    for sheet_name, sheet_cfg in high_cfgs.items():
        if not isinstance(sheet_cfg, dict):
            continue
        if "request_params" not in sheet_cfg and "otp_params" not in sheet_cfg:
            continue
        print(f"[KRX] 신고가({sheet_name}) 조회 중...")
        try:
            raw = fetch_with_retry(client, sheet_cfg, yyyymmdd, f"신고가-{sheet_name}")
            # 실제 신고가 종목만 필터링 (당일 종가 >= 기간 내 최고 종가)
            filtered: list[dict[str, str]] = []
            for row in raw:
                close = as_int(row.get("TDD_CLSPRC", "0"))
                hist_high = as_int(row.get("HGST_ADJ_CLSPRC", "0"))
                if close > 0 and hist_high > 0 and close >= hist_high:
                    filtered.append(row)
            raw = filtered if filtered else raw
            # 누락 필드 보강
            for row in raw:
                code = as_text_code(
                    row.get("ISU_CD", row.get("ISU_SRT_CD", row.get("종목코드", "")))
                )
                if not any(f in row for f in high_fm["market_cap"]):
                    row["시가총액"] = str(mcap_map.get(code, 0))
                if not any(f in row for f in high_fm["market"]):
                    row["시장구분"] = mkt_map.get(code, "")
            rows = parse_high_rows(raw, high_fm, market_cap_unit)
            high_rows_by_sheet[sheet_name] = rows
            print(f"  → {len(rows)}종목")
        except Exception as e:
            print(f"  → 실패: {e}")
        time.sleep(1)

    # 신고가 API 실패 시 시세 데이터 기반 당일 강세 종목으로 폴백
    if not high_rows_by_sheet:
        print("[KRX] 신고가 데이터 없음 → 당일 강세 종목으로 대체")
        candidates: list[HighRow] = []
        for p in price_rows:
            if p.pct <= 0 or p.high <= 0 or p.close <= 0:
                continue
            if p.close < p.high * 0.99:
                continue
            candidates.append(HighRow(
                code=p.code, name=p.name, market=p.market,
                market_cap_eok=as_eok_from_won(p.market_cap),
                pct=p.pct, high_price=p.high,
            ))
        candidates.sort(key=lambda x: x.pct, reverse=True)
        high_rows_by_sheet["당일 강세 종목"] = candidates
        print(f"[KRX] 당일 강세 종목: {len(candidates)}종목")

    # 업종 정보 수집 (DART API)
    industries: dict[str, str] = {}
    if args.dart_config:
        dart_cfg_path = Path(args.dart_config).resolve()
        if dart_cfg_path.exists():
            dart_cfg = load_json(dart_cfg_path)
            dart_api_key = dart_cfg.get("dart", {}).get("api_key", "")
            if dart_api_key:
                unique_codes: set[str] = set()
                for r in top_n(price_rows, lambda x: x.pct, 30):
                    unique_codes.add(r.code)
                for r in top_n(price_rows, lambda x: x.trade_value, 30):
                    unique_codes.add(r.code)
                for r in top_n(
                    [x for x in price_rows if x.market_cap > 0],
                    lambda x: (x.trade_value / x.market_cap), 30,
                ):
                    unique_codes.add(r.code)
                for rows in supply_rows_by_sheet.values():
                    for r in top_n(rows, lambda x: x.net_buy, 10):
                        unique_codes.add(r.code)
                    for r in top_n(rows, lambda x: x.ratio, 10):
                        unique_codes.add(r.code)
                for rows in high_rows_by_sheet.values():
                    for r in rows:
                        unique_codes.add(r.code)
                cache_path = config_path.parent / "industry_cache.json"
                industries = fetch_company_industries(
                    unique_codes, dart_api_key, cache_path=cache_path,
                )
            else:
                print("[업종] dart.api_key가 없어 업종 정보를 건너뜁니다.")
        else:
            print(f"[업종] DART 설정 파일 없음: {dart_cfg_path}")

    out_supply = output_dir / f"{yymmdd}_수급_DATA.xlsx"
    out_price = output_dir / f"{yymmdd}_시세_DATA.xlsx"
    out_high = output_dir / f"{yymmdd}_신고가.xlsx"

    write_supply_workbook(template_supply, out_supply, target_date, supply_rows_by_sheet, industries)
    write_price_workbook(template_price, out_price, target_date, price_rows, industries)
    write_high_workbook(template_high, out_high, high_rows_by_sheet, industries)

    print(f"[완료] {out_supply}")
    print(f"[완료] {out_price}")
    print(f"[완료] {out_high}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"[오류] {exc}", file=sys.stderr)
        raise
